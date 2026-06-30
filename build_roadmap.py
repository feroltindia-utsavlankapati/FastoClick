from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
import copy

wb = Workbook()

# ── Palette ────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
ACCENT_BLUE = "1A56DB"
ACCENT_TEAL = "0E9F6E"
ACCENT_ORG  = "FF5A1F"
ACCENT_PURP = "7E3AF2"
ACCENT_GOLD = "C27803"
LIGHT_GRAY  = "F3F4F6"
MID_GRAY    = "D1D5DB"
WHITE       = "FFFFFF"
P0_RED      = "FEE2E2"
P1_AMBER    = "FEF3C7"
P2_BLUE     = "EFF6FF"
P3_GREEN    = "ECFDF5"

# Priority fills
PFILL = {
    "P0 – Critical": PatternFill("solid", fgColor="FEE2E2"),
    "P1 – High":     PatternFill("solid", fgColor="FEF3C7"),
    "P2 – Medium":   PatternFill("solid", fgColor="EFF6FF"),
    "P3 – Low":      PatternFill("solid", fgColor="ECFDF5"),
}

def hfill(hex_):  return PatternFill("solid", fgColor=hex_)
def bold(sz=10, color=WHITE, italic=False):
    return Font(name="Calibri", bold=True, size=sz, color=color, italic=italic)
def reg(sz=9, color="000000"):
    return Font(name="Calibri", size=sz, color=color)
def wrap_center():
    return Alignment(wrap_text=True, vertical="top", horizontal="center")
def wrap_left():
    return Alignment(wrap_text=True, vertical="top", horizontal="left")
def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    s  = Side(style="thin",   color="D1D5DB")
    sb = Side(style="medium", color="9CA3AF")
    return Border(left=s, right=s, top=s, bottom=sb)

def style_header_row(ws, row, cols, bg, txt=WHITE, sz=10):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = hfill(bg)
        cell.font   = bold(sz, txt)
        cell.alignment = wrap_center()
        cell.border = thin_border()

def style_data_row(ws, row, cols, fill_hex=None, priority=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        if priority and c == 1:
            cell.fill = PFILL.get(priority, hfill(WHITE))
        elif fill_hex:
            cell.fill = hfill(fill_hex)
        else:
            cell.fill = hfill(WHITE) if row % 2 else hfill("F9FAFB")
        cell.font   = reg()
        cell.alignment = wrap_left()
        cell.border = thin_border()

def merge_title(ws, cell, text, bg, fg=WHITE, sz=12, rows=1):
    ws[cell] = text
    ws[cell].fill  = hfill(bg)
    ws[cell].font  = bold(sz, fg)
    ws[cell].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – COVER
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Cover"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 4
for col in "BCDEFGH":
    ws1.column_dimensions[col].width = 22
ws1.row_dimensions[1].height = 8

def cover_block(ws, row, col, text, bg, fg=WHITE, sz=10, bold_=True, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = hfill(bg)
    c.font = Font(name="Calibri", bold=bold_, size=sz, color=fg)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = thin_border()

# Title band
ws1.merge_cells("B2:H4")
ws1["B2"] = "Multi-Agent AI Marketing OS\nProject Implementation Roadmap"
ws1["B2"].fill = hfill(DARK_NAVY)
ws1["B2"].font = Font(name="Calibri", bold=True, size=22, color=WHITE)
ws1["B2"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ws1.row_dimensions[2].height = 28
ws1.row_dimensions[3].height = 28
ws1.row_dimensions[4].height = 20

# Subtitle
ws1.merge_cells("B5:H5")
ws1["B5"] = "FastoClick — Enterprise Agentic AI · Version 2.0 · Confidential"
ws1["B5"].fill = hfill(ACCENT_BLUE)
ws1["B5"].font = bold(11, WHITE, italic=True)
ws1["B5"].alignment = wrap_center()
ws1.row_dimensions[5].height = 18

# Meta grid
meta = [
    ("Project Name",   "Agentic AI Digital Marketing OS (FastoClick)"),
    ("Version",        "2.0 – Complete Re-Architecture"),
    ("Architecture",   "Microservices · Agentic AI Orchestration · Multi-Tenant SaaS"),
    ("Total Duration", "38 Weeks (7 Phases + Continuous Infra)"),
    ("Total Tasks",    "35 Development Tasks across 7 Phases"),
    ("AI Agents",      "10 Autonomous Agents (Strategy, Content, SEO, Ads, Analytics, Email, Social, CRO, Research, Automation)"),
    ("Tech Stack",     "FastAPI · LangGraph · LiteLLM · Qdrant · PostgreSQL · Temporal · Kafka · Next.js 14"),
    ("3rd Party APIs", "Meta · Google Ads · LinkedIn · X · GA4 · GSC · HubSpot · Shopify · Stripe · Mailchimp · Twilio"),
    ("Security",       "JWT RS256 · MFA · RLS · OPA · mTLS · HashiCorp Vault · SAML 2.0 · OIDC · GDPR / SOC2"),
    ("Classification", "Enterprise Technical Documentation"),
    ("Author",         "Senior Technical Project Manager"),
    ("Date",           "June 2026"),
]
ws1.row_dimensions[6].height = 8
row = 7
for label, value in meta:
    ws1.merge_cells(f"B{row}:C{row}")
    ws1.merge_cells(f"D{row}:H{row}")
    cell_l = ws1[f"B{row}"]
    cell_v = ws1[f"D{row}"]
    cell_l.value = label
    cell_v.value = value
    cell_l.fill  = hfill(ACCENT_BLUE)
    cell_l.font  = bold(9, WHITE)
    cell_l.alignment = wrap_center()
    cell_l.border = thin_border()
    cell_v.fill  = hfill(LIGHT_GRAY)
    cell_v.font  = reg(9, DARK_NAVY)
    cell_v.alignment = wrap_left()
    cell_v.border = thin_border()
    ws1.row_dimensions[row].height = 22
    row += 1

# Phase summary table
row += 1
ws1.merge_cells(f"B{row}:H{row}")
ws1[f"B{row}"] = "PHASE OVERVIEW AT A GLANCE"
ws1[f"B{row}"].fill = hfill(DARK_NAVY)
ws1[f"B{row}"].font = bold(11, WHITE)
ws1[f"B{row}"].alignment = wrap_center()
ws1.row_dimensions[row].height = 18
row += 1

phase_headers = ["Phase", "Name", "Duration", "Tasks", "Priority", "Key Milestone", "Primary Agents / Systems"]
phase_colors  = [ACCENT_BLUE, ACCENT_TEAL, ACCENT_PURP, ACCENT_ORG, ACCENT_GOLD, DARK_NAVY, DARK_NAVY]

for ci, hdr in enumerate(phase_headers, start=2):
    c = ws1.cell(row=row, column=ci, value=hdr)
    c.fill = hfill(ACCENT_BLUE)
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws1.row_dimensions[row].height = 20
row += 1

phases_summary = [
    ("Ph 1", "Core Foundation & Auth",         "6 Weeks",  "4",  "P0",  "SaaS Skeleton Live",          "Auth Service · Tenant Manager · API Gateway"),
    ("Ph 2", "Knowledge Base & RAG Layer",      "5 Weeks",  "4",  "P0",  "AI Memory Layer Ready",       "RAG Pipeline · Qdrant · Vector Embeddings"),
    ("Ph 3", "AI Agent Ecosystem",              "6 Weeks",  "5",  "P0",  "10 Agents Operational",       "LangGraph · LiteLLM · DSPy · All 10 Agents"),
    ("Ph 4", "Workflow Engine & Human-in-Loop", "5 Weeks",  "4",  "P1",  "Automation Engine Live",      "Temporal · React Flow · WebSocket Signals"),
    ("Ph 5", "Multi-Platform Integrations",     "6 Weeks",  "6",  "P1",  "11+ Platforms Connected",     "Meta · Google Ads · LinkedIn · GA4 · HubSpot"),
    ("Ph 6", "Analytics, Optimization & UI",    "5 Weeks",  "6",  "P2",  "Self-Improving AI Loops",     "DSPy · MLflow · Recharts · Campaign Analytics"),
    ("Ph 7", "Enterprise Gov. & Scale Infra.",  "5 Weeks",  "6",  "P0",  "Production-Grade Compliance", "OPA · Istio · K8s · Kafka · ClickHouse · Vault"),
]
phase_bgs = [ACCENT_BLUE, ACCENT_TEAL, ACCENT_PURP, ACCENT_ORG, ACCENT_GOLD, "059669", DARK_NAVY]

for pi, pdata in enumerate(phases_summary):
    for ci, val in enumerate(pdata, start=2):
        c = ws1.cell(row=row, column=ci, value=val)
        c.fill = hfill(phase_bgs[pi] if ci == 2 else (LIGHT_GRAY if row % 2 else WHITE))
        c.font = bold(9, WHITE) if ci == 2 else reg(9)
        c.alignment = wrap_center() if ci in (2,3,4,5) else wrap_left()
        c.border = thin_border()
    ws1.row_dimensions[row].height = 22
    row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – MASTER ROADMAP (all phases)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🗺 Master Roadmap")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

col_widths = [6, 26, 14, 8, 46, 32, 28, 30, 28, 22, 10, 10, 28, 32]
col_names  = ["Task ID", "Phase / Task Name", "Category", "Priority",
              "Detailed Description", "Tech Stack", "3rd Party Integrations",
              "FE / BE / DB / Infra Requirements", "Security Controls",
              "Effort", "Sprint", "Owner Role", "Risks & Mitigations", "Success Criteria"]

for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Merge banner
ws2.merge_cells(f"A1:{get_column_letter(len(col_names))}1")
ws2["A1"] = "MULTI-AGENT AI MARKETING OS — COMPLETE IMPLEMENTATION ROADMAP"
ws2["A1"].fill = hfill(DARK_NAVY)
ws2["A1"].font = bold(14, WHITE)
ws2["A1"].alignment = wrap_center()
ws2.row_dimensions[1].height = 28

# Column headers
for i, name in enumerate(col_names, 1):
    c = ws2.cell(row=2, column=i, value=name)
    c.fill = hfill(ACCENT_BLUE)
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws2.row_dimensions[2].height = 32

# ── Full task data ─────────────────────────────────────────────────────────
# Columns: task_id, name, category, priority, description, tech_stack,
#          third_party, requirements, security, effort, sprint, owner, risks, success

ALL_TASKS = [
    # ─── PHASE 1: Core Foundation ────────────────────────────────────────
    ("", "PHASE 1 — CORE FOUNDATION & AUTH", "", "", "Establish the SaaS skeleton: multi-tenant identity, API gateway, centralized audit infrastructure, and CI/CD pipeline. Everything in the system depends on this phase.", "", "", "", "", "6 Weeks", "", "", "", ""),
    ("1.1", "Auth & Identity Service", "Backend", "P0 – Critical",
     "Build a production-grade authentication service supporting Email/Password, Google OAuth2, and Meta OAuth2. Implement JWT (RS256) with refresh token rotation, MFA via TOTP (authenticator app), and role-based access control (RBAC) with granular permission scopes. Includes session invalidation, device fingerprinting, and brute-force protection.",
     "FastAPI 0.111 · Python 3.11 · PostgreSQL 15 · Redis 7 · Alembic",
     "Google OAuth2 (Identity Platform) · Meta OAuth2 · SendGrid (verification emails) · Twilio (SMS OTP fallback)",
     "BE: FastAPI App Factory with blueprints; Auth microservice on port 8001\nDB: Postgres users/sessions/roles schema; Alembic migrations\nInfra: Docker container; Redis for session store & rate limit counters",
     "JWT RS256 w/ 15-min expiry · Refresh token rotation · MFA (TOTP) · Rate limiting 5 req/min · Bcrypt password hashing · CSRF tokens",
     "3 Weeks", "Sprint 1–3", "Senior Backend Engineer",
     "Risk: OAuth token interception → Mitigation: PKCE flow + short TTL tokens\nRisk: Credential stuffing → Mitigation: IP-based rate limiting + CAPTCHA",
     "99.9% login success rate · Auth latency P95 < 200ms · Zero failed MFA bypasses in pentest"),

    ("1.2", "Multi-Tenant Management & Row-Level Security", "Backend", "P0 – Critical",
     "Implement complete tenant isolation using PostgreSQL Row-Level Security policies. Each tenant gets a unique UUID, isolated namespaces, and enforced data boundaries at the DB layer. Build tenant provisioning API, subscription tier management, and tenant-scoped API key generation. Includes soft-delete and tenant archival workflows.",
     "FastAPI · PostgreSQL 15 (RLS) · Alembic · Python 3.11",
     "Stripe (billing/subscription tiers) · Sendgrid (onboarding emails)",
     "BE: Tenant Service microservice; middleware to inject tenant context on every request\nDB: RLS policies on all tables; tenant_id FK enforced by DB triggers\nInfra: Tenant namespace routing in API Gateway",
     "Tenant isolation validation suite · RLS policy enforcement tests · Zero cross-tenant query capability verified",
     "3 Weeks", "Sprint 1–3", "Senior Backend Engineer",
     "Risk: RLS policy bypass via raw SQL → Mitigation: Disable superuser access; all queries via ORM\nRisk: Tenant provisioning race condition → Mitigation: Distributed lock via Redis",
     "Zero cross-tenant data leaks in automated test suite · Tenant provisioning < 2s · 100% RLS policy coverage"),

    ("1.3", "API Gateway & Service Mesh", "Infrastructure", "P0 – Critical",
     "Deploy a centralized API Gateway (Kong or custom FastAPI router) to handle routing, request transformation, auth validation, and load balancing across all microservices. Implement service mesh with mTLS for inter-service communication, service discovery, and distributed tracing via OpenTelemetry.",
     "Kong Gateway / FastAPI · Istio Service Mesh · OpenTelemetry · Jaeger · Python 3.11",
     "None (internal infrastructure)",
     "BE: Gateway routing config; middleware chain for auth → rate-limit → log\nInfra: Istio control plane; mTLS certificates via cert-manager; Jaeger for tracing\nMonitoring: OpenTelemetry collector → Grafana",
     "mTLS for all service-to-service calls · Service-to-service JWT validation · TLS 1.3 only · Certificate rotation automation",
     "2 Weeks", "Sprint 4–5", "DevOps / Platform Engineer",
     "Risk: mTLS certificate expiry → Mitigation: Automated cert-manager rotation\nRisk: Gateway single point of failure → Mitigation: Active-active gateway deployment",
     "Gateway latency overhead < 5ms · 100% mTLS coverage · Distributed traces visible in Jaeger"),

    ("1.4", "CI/CD Pipeline & Developer Infrastructure", "Infrastructure", "P0 – Critical",
     "Set up complete CI/CD pipeline using GitHub Actions. Includes automated testing (unit, integration, e2e), Docker image builds, vulnerability scanning (Trivy), and deployment to staging/production Kubernetes clusters. Includes pre-commit hooks, code quality gates (ruff, mypy), and automated database migration execution.",
     "GitHub Actions · Docker · Kubernetes · Helm · Trivy · pytest · ruff · mypy",
     "GitHub (source control) · DockerHub / GHCR (registry) · SonarQube (code quality optional)",
     "Infra: GitHub Actions workflows; Helm charts for K8s deployments; Dockerfile multi-stage builds\nBE: pytest config; ruff/mypy pre-commit hooks; Alembic migration CI step",
     "Image vulnerability scanning with Trivy · Secrets scanning in CI (no hardcoded keys) · SAST integration",
     "1 Week", "Sprint 5", "DevOps Engineer",
     "Risk: Pipeline flakiness → Mitigation: Retry logic + parallelized test shards\nRisk: Supply chain attack on dependencies → Mitigation: Dependabot + pip-audit",
     "Zero hardcoded secrets in codebase · < 10 min CI run time · 90%+ test coverage gate"),

    # ─── PHASE 2: Knowledge Base & RAG ──────────────────────────────────
    ("", "PHASE 2 — KNOWLEDGE BASE & RAG INTELLIGENCE LAYER", "", "", "Build the foundational memory system for all AI agents. Includes document ingestion pipeline, vector embeddings, semantic search, and long-term memory per tenant. This is the intelligence backbone that powers every AI agent's context.", "", "", "", "", "5 Weeks", "", "", "", ""),

    ("2.1", "Document Ingestion & Processing Pipeline", "AI/Backend", "P0 – Critical",
     "Build an async document processing pipeline supporting PDF, DOCX, TXT, images (OCR), and website scraping. Pipeline stages: Upload → Parse (LangChain loaders) → Chunk (recursive splitter, 512 tokens, 50 overlap) → Clean (PII scan) → Embed (OpenAI text-embedding-3-large or local model) → Store in Qdrant with tenant isolation. Supports batch and real-time ingestion.",
     "FastAPI · LangChain 0.2 · PyMuPDF · python-docx · Tesseract OCR · Celery · Redis",
     "OpenAI API (text-embedding-3-large) · AWS S3 / Cloudflare R2 (file storage) · Unstructured.io (optional advanced parsing)",
     "BE: Ingestion service with Celery task queue; async file upload endpoint\nDB: Qdrant collections per tenant; Postgres metadata store (file name, status, chunk count)\nInfra: Celery workers with Redis broker; S3 bucket for raw file storage",
     "PII scanning on all ingested content (presidio) · File type validation + size limits · Tenant-isolated Qdrant collections · Virus scanning on upload",
     "2 Weeks", "Sprint 6–7", "ML/AI Engineer",
     "Risk: Large file processing timeout → Mitigation: Chunked streaming upload + async Celery task\nRisk: Embedding model rate limits → Mitigation: Token bucket + fallback to local model",
     "P95 ingestion latency < 30s for 10-page PDF · PII detection accuracy > 95% · Zero cross-tenant vector leakage"),

    ("2.2", "Vector Database & Semantic Search Service", "AI/Backend", "P0 – Critical",
     "Deploy and configure Qdrant as the primary vector database with tenant-isolated collections. Build a semantic search API supporting: similarity search, hybrid search (dense + sparse BM25), metadata filtering, and re-ranking via cross-encoder. Implement collection lifecycle management (create, update, delete per tenant).",
     "Qdrant 1.9 · FastAPI · Python 3.11 · sentence-transformers · FlagEmbedding (BGE reranker)",
     "None (self-hosted Qdrant)",
     "BE: Search microservice; hybrid search combining vector + BM25\nDB: Qdrant collections with HNSW index; payload indexing for metadata filters\nInfra: Qdrant docker/K8s deployment; persistent volume for vectors",
     "Tenant collection isolation enforced at API layer · OPA policies for search scope · Qdrant API key rotation",
     "2 Weeks", "Sprint 6–7", "ML/AI Engineer",
     "Risk: Vector index corruption → Mitigation: Qdrant snapshots + WAL enabled\nRisk: Search latency degradation at scale → Mitigation: HNSW tuning + Qdrant sharding",
     "Semantic search P95 < 200ms · Retrieval precision@5 > 80% on benchmark · Zero cross-tenant results"),

    ("2.3", "RAG Pipeline & AI Memory Service", "AI/Backend", "P0 – Critical",
     "Implement the full Retrieval-Augmented Generation pipeline used by all agents. Includes: query expansion, context retrieval from Qdrant, context ranking/compression (LLMLingua or similar), prompt assembly, and LLM call via LiteLLM gateway. Implements short-term (Redis, per-session) and long-term (Qdrant, per-tenant) memory management with memory summarization.",
     "LangChain 0.2 · LiteLLM · Qdrant · Redis · Python 3.11 · LLMLingua (context compression)",
     "OpenAI GPT-4o / Anthropic Claude 3.5 / Gemini 1.5 Pro (via LiteLLM) · Cohere Rerank API",
     "BE: RAG service module; LiteLLM proxy config for multi-provider routing; memory manager class\nDB: Redis for session context (TTL 1h); Qdrant for persistent memory\nInfra: LiteLLM proxy server deployment",
     "Semantic prompt guards (injection detection) · Context window budget enforcement · PII redaction before LLM calls · LLM output logging for audit",
     "2 Weeks", "Sprint 8–9", "ML/AI Engineer",
     "Risk: Context retrieval irrelevance → Mitigation: Query expansion + re-ranking\nRisk: LLM provider outage → Mitigation: LiteLLM auto-fallback chain (GPT-4o → Claude → Gemini)",
     "RAG retrieval P95 < 200ms · RAGAS faithfulness score > 0.85 · LLM fallback tested and verified"),

    ("2.4", "Brand Knowledge Base & Content Guidelines Engine", "AI/Backend", "P1 – High",
     "Build the tenant-specific Brand Knowledge Base where users input: Company Name, Brand Description, Target Audience, Competitors, Brand Tone, Marketing Goals, Industry, and USPs. This data is embedded and stored, forming the foundation for all AI-generated strategies and content. Includes Brand Profile validation and completeness scoring.",
     "FastAPI · Qdrant · LangChain · PostgreSQL · React 18",
     "None (internal system)",
     "BE: Brand profile CRUD API; profile completeness scoring algorithm; auto-embedding on save\nFE: Brand setup wizard UI (multi-step form); completeness progress indicator\nDB: brand_profiles table in Postgres; embedded vectors in Qdrant",
     "Brand data treated as sensitive tenant asset · Access scoped to tenant admins · Versioning for brand profile changes",
     "1 Week", "Sprint 9", "Full-Stack Engineer",
     "Risk: Incomplete brand profiles degrading agent output → Mitigation: Mandatory field validation + completeness gate before agent runs",
     "Brand profile completeness score ≥ 80% before agent activation · Profile embedding latency < 5s"),

    # ─── PHASE 3: AI Agent Ecosystem ────────────────────────────────────
    ("", "PHASE 3 — AI AGENT ECOSYSTEM (10 AUTONOMOUS AGENTS)", "", "", "Build all 10 autonomous AI agents using LangGraph for stateful orchestration. Each agent has a defined role, tool set, memory access, and output schema. Agents communicate through the orchestration layer and all outputs pass through Human-in-the-Loop approval gates.", "", "", "", "", "6 Weeks", "", "", "", ""),

    ("3.1", "Agent Orchestration Framework & Registry", "AI/Backend", "P0 – Critical",
     "Build the central agent orchestration framework using LangGraph. Implement: Agent Registry (self-registration via manifests), Agent Lifecycle Manager (spawn, pause, terminate), Inter-agent Communication Protocol (message queue via Kafka), Agent State Machine (idle → running → waiting → completed → failed), and Tool Registry (declarative tool definitions with access policies).",
     "LangGraph 0.2 · Python 3.11 · Redis · PostgreSQL · Kafka",
     "LiteLLM (LLM gateway for all agents)",
     "BE: Agent Registry service; LangGraph StateGraph definitions; Kafka topics for agent messaging\nDB: agent_registry table; agent_execution_logs; tool_registry\nInfra: Kafka topic provisioning; Redis for agent state caching",
     "CGH Policy Auto-Update on agent registration · OPA tool-access policies · Agent execution audit trail · Sandboxed tool execution",
     "2 Weeks", "Sprint 10–11", "Senior AI Engineer",
     "Risk: Agent state desync → Mitigation: Persistent LangGraph checkpointing to Postgres\nRisk: Runaway agent loops → Mitigation: Max step limits + circuit breaker per agent",
     "Agent registry update < 100ms · Agent state recovery after crash · 100% tool calls logged in audit trail"),

    ("3.2", "Marketing Strategy Planning Agent", "AI Agent", "P0 – Critical",
     "Build the flagship Marketing Strategy Planning Agent. Responsibilities: ingest brand knowledge base context, analyze industry trends, generate comprehensive marketing roadmaps, content pillars, growth strategies, funnel strategies, paid ads recommendations, SEO strategy, and posting schedules. Outputs structured strategy documents stored to DB and presented in UI. Includes audience segmentation and competitor analysis modules.",
     "LangGraph · DSPy · LiteLLM · Python 3.11",
     "OpenAI GPT-4o / Claude 3.5 Sonnet · SerpAPI (competitor research) · SpyFu / SEMrush API (SEO data)",
     "BE: StrategyAgent class extending BaseAgent; DSPy modules for structured output; tool integrations\nFE: Strategy output viewer with rich text rendering; export to PDF/DOCX\nDB: strategies table; strategy_versions for history",
     "Semantic prompt injection guards · Output content safety classifier · Human-in-the-loop approval gate before strategy is published",
     "2 Weeks", "Sprint 10–11", "Senior AI Engineer",
     "Risk: LLM hallucinations in strategy → Mitigation: DSPy structured outputs + confidence scoring + human review gate\nRisk: Generic non-brand-specific output → Mitigation: Mandatory RAG context injection from brand KB",
     "Strategy generation < 60s · Human validator rates 85%+ strategies as brand-aligned · Export to PDF/DOCX functional"),

    ("3.3", "Content Ideas Generator Agent", "AI Agent", "P0 – Critical",
     "Build the Content Ideas Generator Agent that operates on approved strategies. Workflow: Select Strategy → Analyze Context → Generate Ideas → Categorize by Platform → Assign Content Types. Generates: Instagram Posts, Reels scripts, Carousel concepts (with slide-by-slide breakdown), LinkedIn Posts, X/Twitter threads, Blog Topics, YouTube Video Scripts, Shorts/Reels hooks, and CTA suggestions. Applies brand tone and trending topics.",
     "LangGraph · LiteLLM · DSPy · Python 3.11",
     "OpenAI GPT-4o · Google Trends API · BuzzSumo API (trending topics) · Unsplash/Pexels API (image suggestions)",
     "BE: ContentAgent class; platform-specific prompt templates; content categorization logic\nFE: Content idea cards UI with platform badges; one-click add to content calendar\nDB: content_ideas table with platform, type, status fields",
     "Brand tone validation before output · Content safety classifier (hate speech, NSFW filter) · Output tied to approved strategy ID",
     "2 Weeks", "Sprint 12–13", "AI Engineer",
     "Risk: Repetitive/generic content → Mitigation: Diversity constraints in DSPy + trend injection\nRisk: Off-brand tone → Mitigation: Brand tone classifier scoring each output",
     "Generate 20+ content ideas per strategy in < 30s · Brand tone compliance score > 90% · All outputs linked to parent strategy"),

    ("3.4", "SEO Strategy Agent", "AI Agent", "P1 – High",
     "Build an SEO-specialized agent that: researches keywords (volume, difficulty, intent), analyzes competitor rankings, generates content clusters and pillar pages, creates meta title/description templates, identifies technical SEO opportunities, and produces monthly SEO roadmaps. Integrates with Google Search Console for real performance data.",
     "LangGraph · LiteLLM · Python 3.11",
     "SEMrush API / Ahrefs API (keyword research) · Google Search Console API · DataForSEO API · Moz API",
     "BE: SEOAgent class; keyword research tool wrappers; GSC data fetcher\nFE: SEO dashboard with keyword cluster visualization\nDB: seo_keywords table; seo_reports table",
     "API keys stored in HashiCorp Vault · Rate limit handling per API · User data from GSC treated as PII",
     "1 Week", "Sprint 12", "AI Engineer",
     "Risk: Keyword data API costs → Mitigation: Cache results for 24h; implement daily quota limits\nRisk: GSC API quota exhaustion → Mitigation: Request batching + exponential backoff",
     "Keyword research for 50+ terms in < 60s · GSC data sync working · SEO report export functional"),

    ("3.5", "Ads, Analytics, Email, Social, CRO, Research & Automation Agents", "AI Agent", "P1 – High",
     "Build 6 additional specialized agents:\n• Ads Agent: Generate Google/Meta/LinkedIn ad copy variants, targeting recommendations, bid strategy suggestions\n• Analytics Agent: Interpret campaign performance data, identify anomalies, generate insight narratives\n• Email Marketing Agent: Write email sequences, subject line variants, segmentation recommendations\n• Social Media Agent: Platform-specific content adaptation, optimal posting time recommendations, hashtag research\n• CRO Agent: Landing page copy, A/B test hypothesis generation, conversion funnel analysis\n• Research Agent: Competitor analysis, market trend research, audience insights synthesis\n• Automation Agent: Orchestrate multi-step marketing workflows, trigger-based campaign logic",
     "LangGraph · LiteLLM · DSPy · Python 3.11",
     "Google Ads API · Meta Marketing API · LinkedIn Marketing API · Mailchimp/Klaviyo API · Hotjar API · SimilarWeb API",
     "BE: 6 agent classes extending BaseAgent; shared tool library; inter-agent communication via Kafka\nFE: Agent status dashboard; per-agent output viewers\nDB: agent_outputs table partitioned by agent_type",
     "All agents route through CGH governance layer · Tool access policies enforced by OPA · All outputs logged for compliance",
     "2 Weeks", "Sprint 13–15", "AI Engineer (×2)",
     "Risk: Agent coordination deadlock → Mitigation: Timeout + fallback single-agent mode\nRisk: Over-reliance on single LLM provider → Mitigation: LiteLLM multi-provider routing",
     "All 10 agents active in registry · Each agent produces valid structured output · Inter-agent message passing verified"),

    # ─── PHASE 4: Workflow Engine ─────────────────────────────────────────
    ("", "PHASE 4 — WORKFLOW ENGINE & HUMAN-IN-THE-LOOP", "", "", "Deploy Temporal for durable workflow execution. Build the visual drag-and-drop workflow builder, human approval gates, and automated campaign scheduling. This phase enables complex multi-step marketing automation with guaranteed execution.", "", "", "", "", "5 Weeks", "", "", "", ""),

    ("4.1", "Temporal Infrastructure & Workflow Engine", "Backend/Infra", "P0 – Critical",
     "Deploy and configure Temporal server cluster for durable workflow execution. Implement: Workflow Definitions (Python SDK), Activity functions (individual agent calls, API calls, DB operations), Worker pools (dedicated workers per workflow type), Signal handling (human approvals, external triggers), and workflow history persistence. Implement retry policies and compensation logic.",
     "Temporal 1.22 · Python 3.11 · Temporal Python SDK · PostgreSQL (Temporal DB)",
     "None (self-hosted Temporal)",
     "BE: Temporal client config; workflow definitions; activity implementations; worker pool management\nDB: Temporal PostgreSQL backend (separate from app DB); workflow history storage\nInfra: Temporal server K8s deployment; worker pods with auto-scaling",
     "Workflow execution audit trail · Service Vault integration for activity secrets · Worker mTLS communication",
     "2 Weeks", "Sprint 16–17", "Backend Engineer",
     "Risk: Temporal state corruption → Mitigation: Temporal's built-in event sourcing + regular DB backups\nRisk: Worker pool exhaustion → Mitigation: K8s HPA on worker pods based on Temporal queue depth",
     "100% workflow recovery after forced worker restart · Workflow history queryable · Worker auto-scaling verified"),

    ("4.2", "Campaign Scheduling & Publishing Automation", "Backend", "P1 – High",
     "Build the campaign scheduling system that: queues approved content for multi-platform publishing, manages posting schedules per platform (timezone-aware), handles publishing failures with retry logic, tracks publishing status in real-time, and provides a content calendar view. Integrates with all 5 social platform APIs for automated posting.",
     "Temporal · Celery · Redis · FastAPI · Python 3.11",
     "Meta Graph API (Instagram + Facebook posting) · LinkedIn API · Twitter/X API v2 · YouTube Data API",
     "BE: SchedulerService; Temporal CampaignScheduleWorkflow; Celery tasks for each platform\nFE: Content calendar (FullCalendar.js); drag-and-drop scheduling UI; publishing status tracker\nDB: scheduled_posts table; publishing_logs table",
     "Platform API credentials in HashiCorp Vault · Rate limit tracking per platform per tenant · Publishing action audit log",
     "2 Weeks", "Sprint 17–18", "Backend Engineer",
     "Risk: Platform API rate limiting → Mitigation: Per-platform token bucket + queue overflow handling\nRisk: Timezone miscalculation → Mitigation: Store all times in UTC; convert at display layer only",
     "Successful auto-post to all 5 platforms · Schedule conflict detection working · Publishing failure retry < 3 attempts"),

    ("4.3", "Visual Workflow Builder (Drag & Drop)", "Frontend", "P1 – High",
     "Build the visual no-code workflow builder using React Flow. Users can: drag agent nodes onto canvas, connect agents with directional edges, configure each node's inputs/outputs, set conditional branching logic, add human approval gates, define triggers (schedule/webhook/event), and save/version workflows. Real-time collaboration with WebSocket sync.",
     "React Flow 11 · React 18 · Next.js 14 · WebSockets · FastAPI · Zustand",
     "None (custom internal UI)",
     "FE: React Flow canvas with custom node types; node configuration sidebar; workflow save/load\nBE: Workflow definition API; WebSocket for real-time collaboration; Temporal workflow trigger endpoint\nDB: workflows table (JSON definition storage); workflow_versions table",
     "Workflow definitions validated server-side before execution · Human gate nodes are mandatory for all publish actions · WebSocket connections authenticated",
     "2 Weeks", "Sprint 18–19", "Senior Frontend Engineer",
     "Risk: Complex workflow state sync → Mitigation: Operational Transform (OT) or CRDT for collaboration\nRisk: React Flow performance with many nodes → Mitigation: Virtual rendering for large canvases",
     "WebSocket latency < 50ms · Workflow with 10+ nodes renders without lag · Workflow saves and triggers Temporal execution"),

    ("4.4", "Human-in-the-Loop Approval System", "Full-Stack", "P1 – High",
     "Build the Human-in-the-Loop (HITL) approval system. Features: approval queue dashboard (pending strategy reviews, content approvals, publishing approvals), role-based approval routing (editor/manager/admin), approval with inline comments, bulk approve/reject, mobile-responsive approval interface, email/Slack notifications for pending approvals, and Temporal signal integration to resume/cancel workflows.",
     "FastAPI · React 18 · Temporal Signals · Redis · PostgreSQL · WebSockets",
     "Slack API (approval notifications) · SendGrid (email notifications)",
     "BE: ApprovalService; Temporal signal handlers; notification dispatcher\nFE: Approval queue UI; inline comment editor; bulk action controls; mobile responsive\nDB: approval_requests table; approval_comments table; approval_audit_log",
     "All approval actions cryptographically signed and logged · Approval audit trail immutable · Reviewer identity verified via session JWT",
     "1 Week", "Sprint 20", "Full-Stack Engineer",
     "Risk: Approval bottleneck blocking campaigns → Mitigation: SLA alerts; delegate approval feature; auto-escalation",
     "Approval action recorded in audit log with reviewer ID · Temporal workflow resumes within 1s of approval · Slack notification delivered < 30s"),

    # ─── PHASE 5: Integrations ──────────────────────────────────────────
    ("", "PHASE 5 — MULTI-PLATFORM INTEGRATIONS HUB", "", "", "Connect FastoClick to 11+ external platforms across social media, advertising, analytics, CRM, and e-commerce. All credentials managed via HashiCorp Vault. OAuth2 PKCE flows with automatic token refresh.", "", "", "", "", "6 Weeks", "", "", "", ""),

    ("5.1", "Integration Hub & Secrets Vault", "Backend/Infra", "P1 – High",
     "Build the Integration Hub as the central registry for all third-party connections. Features: OAuth2 PKCE flow management, credential storage in HashiCorp Vault with envelope encryption, integration health monitoring, automatic token refresh, integration status dashboard, and webhook receiver infrastructure. Supports 15+ integration types.",
     "FastAPI · HashiCorp Vault · OAuth2 · Celery · PostgreSQL",
     "HashiCorp Vault (secrets management) · None other (integration hub is the enabler)",
     "BE: IntegrationHub service; OAuth2 callback handlers; Vault client for secret CRUD; token refresh scheduler\nDB: integrations table (type, status, tenant_id); integration_events log\nInfra: HashiCorp Vault K8s deployment; AppRole authentication",
     "PKCE OAuth2 flow (no implicit grant) · All secrets envelope-encrypted in Vault · Zero unencrypted credentials in application DB · Secret rotation automation",
     "1 Week", "Sprint 21", "Backend Engineer",
     "Risk: Vault unavailability → Mitigation: Vault HA cluster; cached credentials with short TTL in Redis\nRisk: Token refresh race condition → Mitigation: Distributed lock per integration during refresh",
     "Zero unencrypted secrets in DB verified · Vault HA tested · Token refresh works silently without user re-auth"),

    ("5.2", "Social Media Platform Integrations", "Backend", "P1 – High",
     "Integrate with all major social media platforms:\n• Meta (Facebook + Instagram): Post creation, story publishing, page insights, ad account access via Meta Graph API v19\n• LinkedIn: Company page posts, analytics, follower data via LinkedIn Marketing API\n• X (Twitter): Tweet posting, thread publishing, analytics via Twitter API v2\n• YouTube: Video upload (description/title), channel analytics via YouTube Data API v3\nEach integration includes: OAuth2 connect flow, data sync, webhook receivers for real-time events, and rate limit management.",
     "FastAPI · OAuth2 · Celery · Python 3.11 · httpx",
     "Meta Graph API v19 · LinkedIn Marketing API v2 · Twitter/X API v2 (Basic tier) · YouTube Data API v3",
     "BE: Platform-specific API clients; unified SocialPublishService interface; webhook handlers\nInfra: Celery queues per platform; webhook endpoint with HMAC verification\nDB: social_accounts, social_posts, social_analytics tables",
     "API keys in Vault · HMAC webhook signature verification · Rate limit state tracked in Redis · PII in analytics masked",
     "2 Weeks", "Sprint 21–22", "Backend Engineer",
     "Risk: Meta API policy changes → Mitigation: Version-pinned API clients + deprecation monitoring\nRisk: X API cost increase → Mitigation: Modular client; swap to alternative if needed",
     "Successful OAuth connect for all 4 platforms · Auto-publish tested to each · Analytics data syncing to DB"),

    ("5.3", "Advertising Platform Integrations", "Backend", "P1 – High",
     "Integrate with Google Ads, Meta Ads, and LinkedIn Ads:\n• Google Ads API: Campaign creation, ad group management, keyword bidding, performance data pull, conversion tracking\n• Meta Ads API: Campaign creation, audience targeting, A/B testing, performance insights\n• LinkedIn Campaign Manager API: Sponsored content creation, audience targeting\nIncludes ad performance dashboard aggregating data across all platforms into a unified view.",
     "FastAPI · Google Ads Python Client · Meta Marketing SDK · Python 3.11",
     "Google Ads API v16 · Meta Marketing API v19 · LinkedIn Campaign Manager API",
     "BE: Ads service; campaign sync; performance data normalization layer\nFE: Unified ads performance dashboard with cross-platform comparison\nDB: ad_campaigns, ad_performance tables",
     "Ads API credentials scoped to minimum required permissions · Spend data treated as financial PII · Ad account access logged per action",
     "2 Weeks", "Sprint 22–23", "Backend Engineer",
     "Risk: Google Ads API complexity → Mitigation: Use official Python client library; comprehensive error handling\nRisk: Ad spend data inaccuracy → Mitigation: Daily reconciliation job comparing API vs. platform UI",
     "Campaign create/read/update working for all 3 platforms · Performance data syncing hourly · Cross-platform comparison chart rendering"),

    ("5.4", "Analytics & CRM Integrations", "Backend", "P2 – Medium",
     "Integrate data sources:\n• Google Analytics 4: Traffic, conversion, user behavior via Data API v1\n• Google Search Console: Impressions, clicks, keyword rankings via Search Console API\n• HubSpot CRM: Contact sync, deal pipeline, form submissions via HubSpot API v3\n• Shopify: Order data, product performance, customer data via Shopify Admin API (for e-commerce clients)\n• WordPress: Post performance, traffic data via WP REST API + Google Analytics\nAll data normalized into a unified data model.",
     "FastAPI · Python 3.11 · Celery · SQLAlchemy",
     "Google Analytics 4 Data API · Google Search Console API · HubSpot API v3 · Shopify Admin API 2024-01 · WordPress REST API",
     "BE: Per-platform data fetchers; DataNormalizationLayer with unified schema; scheduled sync jobs\nDB: analytics_raw table (partitioned by source); analytics_normalized view\nInfra: Celery beat scheduler for hourly/daily syncs",
     "PII scanning on all ingested analytics data · Data retention policies enforced · Customer data from Shopify/HubSpot under GDPR scope",
     "2 Weeks", "Sprint 23–24", "Backend Engineer",
     "Risk: GA4 API quota limits → Mitigation: Incremental data fetching; daily quota distribution\nRisk: Schema drift from platform updates → Mitigation: Schema validation tests run daily against API",
     "Data syncing from all 5 sources · Unified normalized schema working · PII scan on ingestion verified"),

    ("5.5", "Email Marketing & Communication Integrations", "Backend", "P2 – Medium",
     "Integrate email and communication platforms:\n• Mailchimp: Audience sync, campaign creation, performance data via Mailchimp API v3\n• Klaviyo: E-commerce email flows, segment sync via Klaviyo API v2024\n• SendGrid: Transactional email delivery for platform notifications\n• Twilio: SMS notifications for HITL alerts and MFA\nBuilds unified email performance dashboard.",
     "FastAPI · Python 3.11 · Celery",
     "Mailchimp API v3 · Klaviyo API v2024 · SendGrid API v3 · Twilio API",
     "BE: Email platform service clients; unified EmailCampaignService interface; sync jobs\nFE: Email performance dashboard\nDB: email_campaigns, email_analytics tables",
     "Email credentials in Vault · Unsubscribe sync mandatory (CAN-SPAM compliance) · Transactional email volume monitored",
     "1 Week", "Sprint 24", "Backend Engineer",
     "Risk: Mailchimp/Klaviyo data sync conflicts → Mitigation: Sync idempotency keys; last-write-wins with timestamp",
     "Email send via SendGrid working · Mailchimp campaign sync functional · Twilio OTP delivery < 10s"),

    ("5.6", "Payments & Subscription Management", "Backend", "P1 – High",
     "Integrate Stripe for platform monetization:\n• Subscription plans (Starter / Pro / Enterprise) with tiered feature access\n• Usage-based billing for AI token consumption\n• Webhook handlers for subscription events (created, updated, cancelled, payment_failed)\n• Customer portal for self-service billing management\n• Invoice generation and dunning management\nTenant subscription status gates feature access via middleware.",
     "FastAPI · Stripe Python SDK · PostgreSQL · Celery",
     "Stripe API (Payments, Subscriptions, Webhooks, Customer Portal)",
     "BE: StripeService; subscription event webhook handlers; usage metering API; entitlement middleware\nFE: Pricing page; subscription management UI; invoice history\nDB: subscriptions, invoices, usage_events tables",
     "Stripe webhook HMAC signature verification · PCI DSS: no card data stored · Subscription access control enforced at middleware layer",
     "1 Week", "Sprint 21", "Backend Engineer",
     "Risk: Failed payment causing service interruption → Mitigation: Grace period (3 days) + dunning emails\nRisk: Usage metering inaccuracy → Mitigation: Idempotent usage event recording",
     "Subscription creation/cancellation webhooks handled · Feature gates working by plan tier · Usage billing tracked per tenant"),

    # ─── PHASE 6: Analytics, Optimization & UI ────────────────────────
    ("", "PHASE 6 — ANALYTICS ENGINE, AI OPTIMIZATION & FRONTEND UI", "", "", "Build the campaign analytics dashboard, AI self-improvement loops, A/B prompt testing, and the complete frontend UI. This phase makes the system intelligent, data-driven, and user-friendly.", "", "", "", "", "5 Weeks", "", "", "", ""),

    ("6.1", "Campaign Analytics Engine & Dashboard", "Full-Stack", "P2 – Medium",
     "Build the unified campaign analytics engine that aggregates performance data from all connected platforms. Features: real-time metrics (Engagement Rate, Reach, Impressions, CTR, Conversions), audience insights, platform comparison, ROI attribution modeling, custom date range filtering, custom report builder, and data export (CSV/PDF). Predictive insights using simple trend forecasting.",
     "FastAPI · ClickHouse · React 18 · Recharts · Next.js 14 · Python 3.11",
     "All connected ad/social/analytics platforms (data already in normalized DB from Phase 5)",
     "BE: AnalyticsAggregationService; ClickHouse queries for performance data; trend forecasting models\nFE: Analytics dashboard with Recharts; custom report builder; PDF export via Puppeteer\nDB: ClickHouse for time-series analytics data; PostgreSQL for report definitions",
     "Aggregate data anonymized where required · Report access scoped to tenant · PII masked in audience insights",
     "2 Weeks", "Sprint 25–26", "Full-Stack Engineer",
     "Risk: Slow analytics queries → Mitigation: Pre-aggregated materialized views in ClickHouse; query result caching\nRisk: Data freshness issues → Mitigation: Display last sync timestamp; real-time sync for key metrics",
     "Dashboard loads in < 2s · All platform metrics displaying correctly · Custom report export working"),

    ("6.2", "AI Performance Feedback Loops & Confidence Calibration", "AI/Backend", "P2 – Medium",
     "Build AI self-improvement infrastructure: performance feedback pipeline (human ratings → DSPy optimizer → improved prompts), confidence calibration engine (track predicted vs. actual accuracy, recalibrate scoring thresholds), model drift detection (alert when output quality degrades), and automated retraining triggers. Implements Expected Calibration Error (ECE) tracking.",
     "DSPy · MLflow · Python 3.11 · FastAPI · Recharts",
     "MLflow Tracking Server (self-hosted) · None other",
     "BE: FeedbackService; DSPy optimization loop; confidence calibration engine; MLflow experiment tracking\nFE: AI performance dashboard (calibration curves, accuracy trends, confidence gauges)\nDB: agent_feedback, prompt_versions, calibration_metrics tables",
     "PII scanning on feedback data · Anomaly alerts for sudden accuracy drops · Prompt versioning with rollback capability",
     "2 Weeks", "Sprint 25–26", "ML/AI Engineer",
     "Risk: DSPy optimization overfitting → Mitigation: Hold-out validation set; gradual prompt rollout\nRisk: Feedback data bias → Mitigation: Reviewer diversity tracking; bias detection alerts",
     "Corr(Confidence, Accuracy) > 0.8 · ECE < 0.05 · Prompt optimization cycle < 24h"),

    ("6.3", "A/B Prompt Testing & Experiment Management", "AI/Backend", "P2 – Medium",
     "Build multi-variant prompt testing infrastructure: create A/B (and multi-arm) prompt experiments, route a percentage of agent calls to variant prompts, track performance metrics per variant, run statistical significance tests (z-test for proportions), automatically promote winning prompts, and provide experiment management UI with experiment history.",
     "Python 3.11 · MLflow · FastAPI · Scipy · React 18",
     "MLflow (experiment tracking)",
     "BE: ExperimentService; traffic splitting router; statistical significance calculator; MLflow integration\nFE: Experiment creation UI; variant comparison dashboard; winner promotion controls\nDB: experiments, variants, experiment_metrics tables",
     "Prompt injection guards on all test variants · Experiment results isolated per tenant · Statistical significance threshold enforced before auto-promote",
     "1 Week", "Sprint 27", "ML/AI Engineer",
     "Risk: Underpowered experiments (insufficient sample size) → Mitigation: Minimum sample size calculator; experiment duration estimator",
     "Stat-significant winner detected in < 48h (simulated) · Auto-promotion of winner working · Experiment history queryable"),

    ("6.4", "Frontend Dashboard — Central Control Panel", "Frontend", "P1 – High",
     "Build the main application dashboard (Central Control Panel) with: Overview metrics widgets, Recent activities feed, Quick-action shortcuts (Generate Strategy, Create Content, Schedule Post), Notification center, System health indicators, Connected platforms status panel, Campaign performance summary, Upcoming scheduled posts preview. Fully responsive (mobile + tablet + desktop). Dark mode support.",
     "Next.js 14 (App Router) · React 18 · Tailwind CSS 3 · Recharts · Zustand · React Query",
     "None (consumes internal APIs)",
     "FE: Dashboard page components; Zustand global state; React Query for API caching; responsive layouts\nBE: Dashboard summary API endpoint (aggregated metrics)\nInfra: Next.js deployed on Vercel or self-hosted K8s",
     "CSP headers · XSS protection · All API calls via authenticated sessions · Dark mode preference stored server-side",
     "1 Week", "Sprint 26", "Senior Frontend Engineer",
     "Risk: Dashboard performance on slow connections → Mitigation: Skeleton loading states; incremental data loading\nRisk: Widget data staleness → Mitigation: React Query with 30s stale time; manual refresh",
     "Dashboard FCP < 1.5s · All widgets functional · Mobile responsive verified on iOS + Android"),

    ("6.5", "Content Calendar & Publishing UI", "Frontend", "P1 – High",
     "Build the Content Calendar interface: monthly/weekly/daily views, drag-and-drop content rescheduling, platform filter toggles (show/hide per platform), content status color coding (Draft/Scheduled/Published/Failed), bulk scheduling actions, content preview per platform format, and direct edit from calendar. Integrates with Phase 4 scheduling engine.",
     "Next.js 14 · React 18 · FullCalendar.js · Tailwind CSS · React DnD",
     "None (internal UI)",
     "FE: Calendar views with FullCalendar; DnD reschedule; content preview modals\nBE: Calendar data API; reschedule endpoint\nDB: scheduled_posts table (already built in Phase 4)",
     "Content actions logged per user · Unauthorized reschedule attempts rejected",
     "1 Week", "Sprint 27", "Frontend Engineer",
     "Risk: Calendar performance with hundreds of items → Mitigation: Virtual scroll + date-range windowing",
     "Drag-and-drop reschedule working · All platform statuses color-coded · Calendar loads < 1s"),

    ("6.6", "Notifications, Settings & User Management UI", "Frontend", "P2 – Medium",
     "Build: Global notification system (in-app + email + Slack alerts), User settings pages (profile, security, connected accounts, notification preferences), Team management UI (invite members, assign roles, manage permissions), API key management page, Billing & subscription management UI, and White-label settings for enterprise tenants.",
     "Next.js 14 · React 18 · Tailwind CSS · React Hook Form · Zod",
     "Slack API (notification preferences) · Stripe Customer Portal (billing UI)",
     "FE: Settings pages; team management UI; notification preferences\nBE: Settings API; team CRUD; notification dispatch service\nDB: notification_preferences, team_members tables",
     "Team member invite links time-limited · Role changes logged in audit trail · API keys displayed only once at creation",
     "1 Week", "Sprint 28", "Frontend Engineer",
     "Risk: Settings changes not persisting → Mitigation: Optimistic UI with server reconciliation",
     "Team invite flow end-to-end working · Notification preferences respected · API key generation secure"),

    # ─── PHASE 7: Enterprise Governance & Scale ─────────────────────────
    ("", "PHASE 7 — ENTERPRISE GOVERNANCE, COMPLIANCE & SCALE INFRASTRUCTURE", "", "", "Production-grade governance, compliance (SOC2, GDPR), and infrastructure scaling for enterprise deployment. This phase enables enterprise sales and ensures regulatory compliance.", "", "", "", "", "5 Weeks", "", "", "", ""),

    ("7.1", "Central Governance & Compliance Handler (CGH)", "Backend/Security", "P0 – Critical",
     "Build the full CGH (Central Governance & Compliance Handler) system: Policy Engine (OPA for declarative policy-as-code), Anomaly Detector (ML-based behavioral analysis for unusual agent actions), Action Controller (approve/block/quarantine actions in real-time), Compliance Monitor (continuous compliance posture scoring), and Incident Response Automation (auto-remediation for common violations).",
     "Open Policy Agent (OPA) · Python 3.11 · FastAPI · Kafka · PostgreSQL",
     "AWS S3 / Cloudflare R2 (audit log archival)",
     "BE: CGH service; OPA policy evaluation endpoint; anomaly detection ML model; action controller middleware\nDB: compliance_policies, policy_violations, compliance_scores tables\nInfra: OPA server deployment; Kafka audit topic",
     "100% policy enforcement coverage · Envelope encryption on audit logs · Immutable audit log design (append-only) · GDPR Article 30 record keeping",
     "2 Weeks", "Sprint 29–30", "Security Engineer",
     "Risk: OPA policy false positives blocking legitimate actions → Mitigation: Policy dry-run mode; gradual rollout with shadow enforcement first\nRisk: Audit log tampering → Mitigation: Append-only log with hash chaining; S3 Object Lock",
     "100% policy enforcement · Zero audit log tampering capability · Compliance score dashboard operational"),

    ("7.2", "SSO, Enterprise Auth & Audit Signing Service", "Backend/Security", "P1 – High",
     "Implement enterprise-grade authentication: SAML 2.0 IdP integration (Okta, Azure AD, Google Workspace), OIDC federation, Just-In-Time (JIT) provisioning of users from IdP, cryptographic audit log signing (RSA-2048), immutable signed audit log viewer with non-repudiation guarantees, and session management for enterprise SSO (SCIM 2.0 for user provisioning sync).",
     "Python 3.11 · FastAPI · python3-saml · OIDC · Cryptography lib",
     "Okta API (SSO integration) · Azure Active Directory (SAML/OIDC) · Google Workspace SSO",
     "BE: SAML/OIDC handlers; JIT provisioning logic; audit signer service (RSA signing of log entries)\nFE: Audit log viewer with signature verification display; SSO configuration UI for admins\nDB: sso_configurations, signed_audit_logs tables",
     "Signed, immutable audit logs · Key management via HashiCorp Vault · SCIM provisioning · Session hijacking protection",
     "2 Weeks", "Sprint 29–30", "Security Engineer",
     "Risk: SSO misconfiguration locking out admins → Mitigation: Local fallback admin account always available\nRisk: Signing key compromise → Mitigation: Key rotation schedule + Vault-based key storage",
     "SSO login working with Okta and Azure AD · Audit log signatures verified · SCIM provisioning sync tested"),

    ("7.3", "GDPR & SOC2 Compliance Tooling", "Backend", "P1 – High",
     "Build compliance automation: GDPR tools (data subject access request automation, right-to-erasure workflows, consent management, data processing records Article 30), SOC2 evidence collection automation (control testing, evidence packaging), Data retention policy engine (auto-delete after configured periods), PII discovery and redaction service, and Privacy-by-design audit checklist.",
     "Python 3.11 · FastAPI · Celery · PostgreSQL · Presidio (Microsoft PII)",
     "OneTrust API (optional DPA management) · None required",
     "BE: ComplianceService; DSAR automation; retention policy engine; PII redaction pipeline\nFE: Compliance dashboard (SOC2 control status, GDPR data map)\nDB: data_processing_records, retention_policies, dsar_requests tables",
     "Automated PII redaction in exports · DSAR response within 30-day SLA · Data retention triggers tested",
     "1 Week", "Sprint 31", "Backend Engineer",
     "Risk: Regulatory scope changes (AI Act) → Mitigation: Modular compliance framework; legal review process",
     "Mock SOC2 audit passed · DSAR automation completes in < 24h · GDPR data map complete"),

    ("7.4", "Multi-Region Kubernetes & High-Availability Infrastructure", "Infrastructure", "P2 – Medium",
     "Deploy multi-region Kubernetes clusters for production: primary region (e.g., us-east-1) + secondary (eu-west-1 for GDPR), global load balancer (Cloudflare / AWS Global Accelerator), database read replicas in each region, cross-region DR procedures, automated failover testing, and RTO < 15min / RPO < 5min targets. Includes infrastructure-as-code via Terraform.",
     "Kubernetes 1.29 · Istio · Terraform · Helm · Cloudflare / AWS",
     "AWS EKS / GKE / DigitalOcean K8s · Cloudflare (CDN + DDoS protection) · AWS RDS Multi-AZ",
     "Infra: Terraform modules for K8s cluster; multi-region DNS; DB read replicas; cross-region Kafka replication\nMonitoring: Grafana + Prometheus multi-cluster; PagerDuty alerting\nDB: PostgreSQL streaming replication; Qdrant distributed cluster",
     "Multi-region DR procedures documented · Regular failover drills scheduled · Encryption in transit and at rest",
     "2 Weeks", "Ongoing from Sprint 29", "DevOps / Platform Engineer",
     "Risk: Region sync lag → Mitigation: Async replication with eventual consistency; strong consistency only where required\nRisk: Terraform state corruption → Mitigation: Remote state in S3 with locking via DynamoDB",
     "Zero-downtime failover verified in DR drill · RTO < 15min demonstrated · Multi-region traffic routing working"),

    ("7.5", "High-Volume Data Layer (Kafka + ClickHouse)", "Infrastructure", "P2 – Medium",
     "Scale the data infrastructure for high-volume operations: Kafka multi-broker cluster for event streaming (audit events, analytics events, agent messages), ClickHouse cluster with sharding for time-series analytics (campaign performance, agent metrics, user events), data pipelines from Kafka → ClickHouse, and real-time dashboards on ClickHouse data. Targets 1M+ events/day.",
     "Apache Kafka 3.6 · ClickHouse 24.x · Python 3.11 · Kafka Connect",
     "None (self-hosted)",
     "Infra: Kafka K8s deployment (3+ brokers); ClickHouse K8s cluster (3 shards × 2 replicas); Kafka Connect pipeline\nBE: Kafka producers in all services; ClickHouse query optimization\nDB: ClickHouse schemas for analytics; Kafka topic partitioning strategy",
     "Audit trail encryption on Kafka · ClickHouse access restricted to analytics service · Data retention policies enforced in ClickHouse",
     "2 Weeks", "Ongoing", "Data/Platform Engineer",
     "Risk: Kafka consumer lag → Mitigation: Consumer group monitoring + auto-scaling consumers\nRisk: ClickHouse disk cost → Mitigation: TTL policies on old data; cold storage tiering",
     "ClickHouse query latency < 1s at 100M rows · Kafka consumer lag < 10s · 1M events/day ingestion tested"),

    ("7.6", "Performance Optimization & Observability Stack", "Infrastructure", "P2 – Medium",
     "Complete observability and performance optimization: Full Grafana + Prometheus monitoring stack (200+ metrics across all services), centralized logging (ELK stack or Grafana Loki), distributed tracing (Jaeger/Tempo), performance profiling and optimization pass across all services (target P95 Agent latency < 30s), SLO/SLA definition and alerting, runbook automation for common incidents, and load testing suite (k6/Locust).",
     "Grafana · Prometheus · Jaeger / Tempo · ELK / Loki · k6 · Python 3.11",
     "PagerDuty (on-call alerting) · Sentry (error tracking) · Datadog (optional premium observability)",
     "Infra: Monitoring stack K8s deployment; alert rules; SLO dashboards\nBE: OpenTelemetry instrumentation on all services; structured logging\nOps: Runbooks for common incidents; on-call rotation setup",
     "mTLS everywhere enforced in service mesh · Log access restricted by role · PagerDuty alerts for P0 incidents < 1min",
     "1 Week", "Sprint 32–33", "DevOps Engineer",
     "Risk: Alert fatigue → Mitigation: Alert noise reduction; severity-based routing; grouping\nRisk: Observability data cost → Mitigation: Metric cardinality limits; log sampling for non-error traces",
     "P95 agent latency < 30s · 200+ metrics visible in Grafana · Load test simulates 100 concurrent users successfully"),
]

# Write tasks to sheet
current_row = 3
for task in ALL_TASKS:
    task_id, name, category, priority, desc, tech, third_party, reqs, security, effort, sprint, owner, risks, success = task
    is_phase_header = task_id == "" and name.startswith("PHASE")

    values = [task_id, name, category, priority, desc, tech, third_party, reqs, security, effort, sprint, owner, risks, success]
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=current_row, column=col_idx, value=val)
        cell.border = thin_border()
        if is_phase_header:
            cell.fill   = hfill(DARK_NAVY)
            cell.font   = bold(10, WHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left" if col_idx > 1 else "center")
            ws2.row_dimensions[current_row].height = 24
        else:
            if priority:
                cell.fill = PFILL.get(priority, hfill(WHITE if current_row % 2 else "F9FAFB"))
            else:
                cell.fill = hfill(WHITE if current_row % 2 else "F9FAFB")
            cell.font = bold(9, WHITE) if col_idx == 2 and not is_phase_header else reg(9)
            if col_idx == 2 and not is_phase_header:
                cell.fill = hfill(ACCENT_BLUE)
            cell.alignment = wrap_left()
            ws2.row_dimensions[current_row].height = 80

    current_row += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – 3RD PARTY INTEGRATIONS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🔌 3rd Party Integrations")
ws3.sheet_view.showGridLines = False

int_cols = ["Category", "Service / API", "Provider", "API Version", "Purpose in System",
            "Auth Method", "Rate Limits", "Estimated Cost", "Phase Introduced", "Fallback / Alternative"]
int_widths = [18, 22, 18, 14, 42, 18, 22, 20, 16, 22]
for i, w in enumerate(int_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells(f"A1:{get_column_letter(len(int_cols))}1")
ws3["A1"] = "THIRD-PARTY INTEGRATIONS & EXTERNAL RESOURCES — COMPLETE REGISTRY"
ws3["A1"].fill = hfill(DARK_NAVY)
ws3["A1"].font = bold(13, WHITE)
ws3["A1"].alignment = wrap_center()
ws3.row_dimensions[1].height = 26

for i, h in enumerate(int_cols, 1):
    c = ws3.cell(row=2, column=i, value=h)
    c.fill  = hfill(ACCENT_TEAL)
    c.font  = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws3.row_dimensions[2].height = 28

INTEGRATIONS = [
    # Category, Service, Provider, Version, Purpose, Auth, Rate Limits, Cost, Phase, Fallback
    ("🤖 AI / LLM", "GPT-4o", "OpenAI", "API v1", "Primary LLM for strategy, content generation, and complex reasoning tasks", "API Key (Bearer)", "500 RPM / 10K TPM (Tier 2)", "~$15–$60 / 1M tokens", "Phase 3", "Claude 3.5 Sonnet (Anthropic)"),
    ("🤖 AI / LLM", "Claude 3.5 Sonnet", "Anthropic", "API v1", "Fallback LLM; used for long-context tasks and code generation", "API Key (Bearer)", "50 RPM / 100K TPM", "~$3–$15 / 1M tokens", "Phase 3", "GPT-4o"),
    ("🤖 AI / LLM", "Gemini 1.5 Pro", "Google", "Generative AI API v1beta", "Tertiary LLM fallback; used for multimodal tasks with image input", "API Key / Service Account", "60 RPM", "~$3.50 / 1M tokens", "Phase 3", "GPT-4o Vision"),
    ("🤖 AI / LLM", "LiteLLM Proxy", "LiteLLM (OSS)", "Latest", "Multi-provider LLM gateway with failover, cost tracking, and rate limiting", "Internal Service", "Unlimited (proxy)", "Free (OSS)", "Phase 3", "Direct API calls"),
    ("🤖 AI / LLM", "text-embedding-3-large", "OpenAI", "API v1", "Document and content embeddings for RAG and semantic search", "API Key (Bearer)", "3000 RPM", "~$0.13 / 1M tokens", "Phase 2", "BGE-M3 (local, free)"),
    ("🤖 AI / LLM", "Cohere Rerank API", "Cohere", "API v1", "Re-ranking retrieved RAG context for relevance improvement", "API Key", "100 RPM (free tier)", "Free tier / $1 per 1K searches", "Phase 2", "BGE Reranker (local)"),
    ("📱 Social Media", "Meta Graph API", "Meta (Facebook)", "v19.0", "Facebook/Instagram post publishing, page insights, story posting, ad account", "OAuth2 (User + Page tokens)", "200 calls/hour per user token", "Free (platform API)", "Phase 5", "Buffer API (fallback publishing)"),
    ("📱 Social Media", "Instagram Basic Display API", "Meta", "v19.0", "Instagram media publishing and account analytics", "OAuth2", "200 calls/hour", "Free", "Phase 5", "Meta Graph API (same)"),
    ("📱 Social Media", "LinkedIn Marketing API", "LinkedIn", "v2", "Company page posts, sponsored content, follower analytics", "OAuth2 (3-legged)", "500 calls/day per app", "Free (approval required)", "Phase 5", "Hootsuite API"),
    ("📱 Social Media", "Twitter / X API v2", "X Corp", "v2 (Basic tier)", "Tweet posting, thread creation, analytics, follower data", "OAuth2 / Bearer Token", "500K tweets/month (Basic)", "$100/month (Basic tier)", "Phase 5", "SocialBee API"),
    ("📱 Social Media", "YouTube Data API v3", "Google", "v3", "Channel analytics, video descriptions, upload metadata", "OAuth2 / API Key", "10,000 units/day", "Free (quota-based)", "Phase 5", "Manual posting fallback"),
    ("📢 Advertising", "Google Ads API", "Google", "v16", "Campaign management, keyword bidding, ad copy, conversion tracking, performance data", "OAuth2 (Manager Account)", "15,000 operations/day", "Free (Google Partner required)", "Phase 5", "Manual Ads management"),
    ("📢 Advertising", "Meta Marketing API", "Meta", "v19.0", "Facebook/Instagram ad campaigns, audience targeting, A/B tests, insights", "OAuth2 (Ad Account)", "200 calls/hour", "Free (platform API)", "Phase 5", "Google Ads only"),
    ("📢 Advertising", "LinkedIn Campaign Manager API", "LinkedIn", "v2", "Sponsored content campaigns, audience targeting, performance insights", "OAuth2", "500 calls/day", "Free (approval needed)", "Phase 5", "Manual LinkedIn Ads"),
    ("📊 Analytics", "Google Analytics 4 Data API", "Google", "Data API v1", "Website traffic, user behavior, conversions, audience data", "OAuth2 / Service Account", "200K requests/day", "Free", "Phase 5", "Matomo (self-hosted)"),
    ("📊 Analytics", "Google Search Console API", "Google", "v3", "Search impressions, clicks, keyword rankings, indexing status", "OAuth2 / Service Account", "200 queries/100 seconds", "Free", "Phase 5", "Ahrefs API"),
    ("📊 Analytics", "SEMrush API", "SEMrush", "API v3", "Keyword research, competitor ranking data, backlink analysis, content gap", "API Key", "10 requests/second", "From $100/month", "Phase 3", "Ahrefs API / DataForSEO"),
    ("📊 Analytics", "BuzzSumo API", "BuzzSumo", "API v1", "Trending content topics, social share counts, influencer data", "API Key", "10K requests/month", "From $199/month", "Phase 3", "Google Trends API (free)"),
    ("📊 Analytics", "SerpAPI", "SerpAPI", "API v2", "SERP data for competitor research and keyword position tracking", "API Key", "100 searches/month (free)", "Free / $50+ per month", "Phase 3", "DataForSEO SERP API"),
    ("🤝 CRM", "HubSpot API", "HubSpot", "v3 (CRM API)", "Contact sync, deal pipeline, form submissions, marketing email data", "OAuth2 / Private App Token", "100 requests/10 seconds", "Free CRM / API from $45/month", "Phase 5", "Salesforce API"),
    ("🛒 E-Commerce", "Shopify Admin API", "Shopify", "2024-01", "Order data, product performance, customer segments, sales analytics", "OAuth2 / Admin API Key", "40 requests/app/second (Plus: 80)", "Free with Shopify subscription", "Phase 5", "WooCommerce REST API"),
    ("📧 Email", "Mailchimp API", "Mailchimp (Intuit)", "API v3", "Audience sync, email campaign creation, performance reporting, automation", "OAuth2 / API Key", "10 requests/second", "Free (500 contacts) / From $13/month", "Phase 5", "Klaviyo API"),
    ("📧 Email", "Klaviyo API", "Klaviyo", "API v2024-02", "E-commerce email flows, behavioral segments, SMS + email unification", "Private API Key", "75 requests/second", "Free (250 contacts) / Usage-based", "Phase 5", "Mailchimp API"),
    ("📧 Email", "SendGrid API", "Twilio SendGrid", "API v3", "Transactional email delivery (verification, notifications, invoices)", "API Key", "600 requests/minute", "Free (100/day) / From $19.95/month", "Phase 1", "AWS SES / Postmark"),
    ("📱 SMS/Voice", "Twilio API", "Twilio", "REST API 2010-04-01", "SMS OTP for MFA fallback, HITL approval SMS alerts", "Account SID + Auth Token", "100 requests/second", "~$0.0079/SMS", "Phase 1", "AWS SNS / Vonage"),
    ("💳 Payments", "Stripe API", "Stripe", "API v1 (2024-04-10)", "SaaS subscription billing, usage-based metering, customer portal, invoicing", "Secret Key / Webhook Sig", "100 reads/sec, 100 writes/sec", "2.9% + 30¢ per transaction", "Phase 5", "Paddle / Lemonsqueezy"),
    ("🔐 Auth", "Google OAuth2", "Google", "OAuth 2.0 / OIDC", "Social login for end users; SSO for Google Workspace enterprise clients", "OAuth2 Client Credentials", "Google rate limits", "Free", "Phase 1", "Microsoft OIDC"),
    ("🔐 Auth", "Meta OAuth2", "Meta", "OAuth 2.0", "Social login via Facebook account", "OAuth2 Client Credentials", "Meta app limits", "Free", "Phase 1", "Google OAuth2"),
    ("🔐 Auth", "Okta SSO", "Okta", "SAML 2.0 / OIDC", "Enterprise SSO integration; SCIM 2.0 user provisioning for enterprise clients", "SAML / OIDC / SCIM", "Enterprise plan limits", "From $6/user/month", "Phase 7", "Azure AD / Auth0"),
    ("🔐 Auth", "Azure Active Directory", "Microsoft", "SAML 2.0 / OIDC", "Enterprise SSO for Microsoft 365 clients; Azure AD SCIM provisioning", "SAML 2.0 / OAuth2", "Microsoft rate limits", "Part of Azure subscription", "Phase 7", "Okta"),
    ("☁️ Infrastructure", "AWS S3 / Cloudflare R2", "AWS / Cloudflare", "Latest", "File storage for uploaded documents, exported reports, audit log archives", "IAM Role / R2 API Token", "Unlimited (cost-based)", "S3: ~$0.023/GB; R2: $0.015/GB (no egress fees)", "Phase 2", "Google Cloud Storage"),
    ("☁️ Infrastructure", "HashiCorp Vault", "HashiCorp", "1.15 (OSS)", "Secret management: API keys, OAuth tokens, DB passwords, encryption keys", "AppRole / K8s Auth", "Internal service", "Free (OSS self-hosted)", "Phase 5", "AWS Secrets Manager"),
    ("☁️ Infrastructure", "Cloudflare (CDN + DNS)", "Cloudflare", "API v4", "CDN, DDoS protection, DNS management, WAF, bot detection", "API Token", "1200 requests/5min", "Free / Pro from $20/month", "Phase 7", "AWS CloudFront"),
    ("📈 Monitoring", "Sentry", "Sentry", "SDK 2.x", "Application error tracking, performance monitoring, source map upload", "DSN Token", "5K errors/month (free)", "Free / From $26/month", "Phase 7", "Rollbar / Bugsnag"),
    ("📈 Monitoring", "PagerDuty", "PagerDuty", "REST API v2", "On-call alerting, incident management, escalation policies", "API Key", "Enterprise plan", "Free (5 users) / From $21/user/month", "Phase 7", "OpsGenie / VictorOps"),
    ("🖼 Media", "Unsplash API", "Unsplash", "API v1", "Free high-quality images for content suggestions and social post templates", "API Key", "50 requests/hour (demo)", "Free (with attribution)", "Phase 3", "Pexels API"),
    ("🖼 Media", "Pexels API", "Pexels", "API v1", "Stock photos and videos for content generation context", "API Key", "200 requests/hour", "Free", "Phase 3", "Unsplash API"),
    ("📊 SEO/Research", "DataForSEO API", "DataForSEO", "REST API v3", "Keyword data, SERP results, backlink data, domain analytics (cost-effective SEMrush alt)", "Login/Password (Basic Auth)", "Unlimited (credit-based)", "Pay-per-use from $50", "Phase 3", "SEMrush API"),
    ("🧠 AI Tools", "MLflow Tracking", "MLflow (OSS)", "2.x (self-hosted)", "Experiment tracking for A/B prompt testing, model performance, run comparisons", "Internal service", "Unlimited (self-hosted)", "Free (OSS)", "Phase 6", "Weights & Biases"),
    ("🧠 AI Tools", "Microsoft Presidio", "Microsoft (OSS)", "2.x", "PII detection and anonymization in text data for GDPR compliance", "Python library", "Local processing", "Free (OSS)", "Phase 2", "AWS Comprehend (PII)"),
]

cat_colors = {
    "🤖 AI / LLM": ACCENT_PURP,
    "📱 Social Media": ACCENT_TEAL,
    "📢 Advertising": ACCENT_ORG,
    "📊 Analytics": ACCENT_BLUE,
    "🤝 CRM": "0E9F6E",
    "🛒 E-Commerce": "C27803",
    "📧 Email": "1C64F2",
    "📱 SMS/Voice": "7E3AF2",
    "💳 Payments": "057A55",
    "🔐 Auth": DARK_NAVY,
    "☁️ Infrastructure": "374151",
    "📈 Monitoring": ACCENT_ORG,
    "🖼 Media": "6B7280",
    "📊 SEO/Research": ACCENT_PURP,
    "🧠 AI Tools": "1E40AF",
}

for row_i, row_data in enumerate(INTEGRATIONS, start=3):
    cat = row_data[0]
    bg = cat_colors.get(cat, ACCENT_BLUE)
    for col_i, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        if col_i == 1:
            cell.fill = hfill(bg)
            cell.font = bold(8, WHITE)
            cell.alignment = wrap_center()
        else:
            cell.fill = hfill(WHITE if row_i % 2 else "F0FDF4")
            cell.font = reg(9)
            cell.alignment = wrap_left()
    ws3.row_dimensions[row_i].height = 36

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – SPRINT PLAN
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📅 Sprint Plan")
ws4.sheet_view.showGridLines = False

sp_cols = ["Sprint", "Weeks", "Phase", "Tasks", "Deliverables", "Team", "Exit Criteria"]
sp_widths = [10, 12, 24, 20, 48, 28, 38]
for i, w in enumerate(sp_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells(f"A1:{get_column_letter(len(sp_cols))}1")
ws4["A1"] = "SPRINT-BY-SPRINT DELIVERY PLAN — 33 SPRINTS · 38 WEEKS"
ws4["A1"].fill = hfill(DARK_NAVY)
ws4["A1"].font = bold(13, WHITE)
ws4["A1"].alignment = wrap_center()
ws4.row_dimensions[1].height = 26

for i, h in enumerate(sp_cols, 1):
    c = ws4.cell(row=2, column=i, value=h)
    c.fill = hfill(ACCENT_PURP)
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws4.row_dimensions[2].height = 26

SPRINTS = [
    ("S1–S3",  "Wk 1–3",  "Ph 1: Core Foundation",      "1.1, 1.2",    "Auth service live, OAuth2 working, tenant isolation DB schema, RLS policies enforced",                                 "2× Senior BE Engineer",             "99.9% login success · Zero cross-tenant leaks in test suite"),
    ("S4–S5",  "Wk 4–5",  "Ph 1: Core Foundation",      "1.3, 1.4",    "API Gateway routing all services, mTLS active, CI/CD pipeline live, Docker builds automated",                         "DevOps Engineer + BE Engineer",      "Gateway overhead < 5ms · CI run < 10min · Trivy zero critical CVEs"),
    ("S6–S7",  "Wk 6–7",  "Ph 2: Knowledge Base",       "2.1, 2.2",    "Document ingestion pipeline (PDF/DOCX/TXT), Qdrant deployed, semantic search API live",                               "ML/AI Engineer",                    "P95 ingestion < 30s · Semantic search P95 < 200ms"),
    ("S8–S9",  "Wk 8–9",  "Ph 2: Knowledge Base",       "2.3, 2.4",    "RAG pipeline with multi-provider LLM, LiteLLM gateway, brand profile system, embedding pipeline",                    "ML/AI Engineer",                    "RAGAS faithfulness > 0.85 · Brand profiles creatable in UI"),
    ("S10–S11","Wk 10–11", "Ph 3: AI Agents",            "3.1, 3.2",    "Agent Registry, LangGraph orchestration framework, Marketing Strategy Agent generating structured strategies",       "Senior AI Engineer",                "Agent state recovery working · Strategy generation < 60s"),
    ("S12–S13","Wk 12–13", "Ph 3: AI Agents",            "3.3, 3.4",    "Content Ideas Agent live, SEO Agent with keyword research, brand tone compliance scoring",                           "AI Engineer",                       "20+ content ideas per strategy · SEO keyword research functional"),
    ("S14–S15","Wk 14–15", "Ph 3: AI Agents",            "3.5",         "All 10 agents in registry: Ads, Analytics, Email, Social, CRO, Research, Automation agents built and tested",       "AI Engineer (×2)",                  "All 10 agents producing valid structured outputs"),
    ("S16–S17","Wk 16–17", "Ph 4: Workflow Engine",      "4.1",         "Temporal cluster deployed, workflow definitions, worker pools, crash recovery tested",                               "Backend Engineer",                  "100% workflow recovery after crash · History queryable"),
    ("S17–S18","Wk 17–18", "Ph 4: Workflow Engine",      "4.2",         "Campaign scheduling engine, auto-publish to all 5 social platforms, content calendar backend",                       "Backend Engineer",                  "Auto-post working on all 5 platforms · Schedule conflict detection"),
    ("S18–S19","Wk 18–19", "Ph 4: Workflow Engine",      "4.3",         "Visual drag-and-drop workflow builder (React Flow), node canvas, WebSocket real-time sync",                          "Senior Frontend Engineer",          "WebSocket latency < 50ms · 10-node workflow builds and triggers"),
    ("S20",    "Wk 20",    "Ph 4: Workflow Engine",      "4.4",         "Human-in-the-loop approval queue, role-based approval routing, Temporal signal integration, Slack notifications",    "Full-Stack Engineer",               "Approval audit logged · Temporal resumes within 1s of approval"),
    ("S21",    "Wk 21",    "Ph 5: Integrations",         "5.1, 5.6",    "Integration Hub with Vault, Stripe subscription billing, OAuth2 connect flows, tenant feature gates",               "Backend Engineer",                  "Zero unencrypted secrets · Stripe webhooks handled · Feature gates live"),
    ("S21–S22","Wk 21–22", "Ph 5: Integrations",         "5.2",         "Social platform OAuth2 connections: Meta, LinkedIn, X, YouTube — publish and analytics sync",                        "Backend Engineer",                  "OAuth connect + auto-publish verified on all 4 social platforms"),
    ("S22–S23","Wk 22–23", "Ph 5: Integrations",         "5.3",         "Google Ads, Meta Ads, LinkedIn Ads — campaign management and performance data sync",                                  "Backend Engineer",                  "Campaign CRUD working · Performance data syncing hourly"),
    ("S23–S24","Wk 23–24", "Ph 5: Integrations",         "5.4, 5.5",    "GA4, GSC, HubSpot, Shopify integrations; email integrations: Mailchimp, Klaviyo, SendGrid, Twilio",                  "Backend Engineer",                  "All 9 sources syncing · Unified data model working"),
    ("S25–S26","Wk 25–26", "Ph 6: Analytics & UI",       "6.1, 6.2",    "Campaign analytics dashboard (ClickHouse), AI feedback loops, DSPy optimization, confidence calibration",            "Full-Stack + ML/AI Engineer",       "Dashboard < 2s · Corr(Conf, Acc) > 0.8 · ECE < 0.05"),
    ("S26–S27","Wk 26–27", "Ph 6: Analytics & UI",       "6.3, 6.4, 6.5", "A/B prompt testing, main dashboard UI, content calendar UI with drag-and-drop",                                  "ML/AI Engineer + Frontend Engineer", "A/B test winner in 48h · Dashboard FCP < 1.5s · Calendar DnD working"),
    ("S28",    "Wk 28",    "Ph 6: Analytics & UI",       "6.6",         "Notifications center, settings pages, team management UI, billing UI, API key management",                           "Frontend Engineer",                 "Team invite flow working · Notification preferences respected"),
    ("S29–S30","Wk 29–30", "Ph 7: Enterprise Gov.",      "7.1, 7.2",    "CGH full implementation (OPA), SSO (SAML/OIDC with Okta + Azure AD), signed immutable audit logs",                  "Security Engineer",                 "100% policy enforcement · SSO with 2 IdPs · Audit log signing verified"),
    ("S31",    "Wk 31",    "Ph 7: Enterprise Gov.",      "7.3",         "GDPR tools (DSAR automation, PII redaction), SOC2 evidence collection, data retention engine",                       "Backend Engineer",                  "Mock SOC2 audit passed · DSAR automation < 24h"),
    ("S32–S33","Wk 32–33", "Ph 7: Scale Infra.",         "7.4, 7.5, 7.6", "Multi-region K8s, Kafka + ClickHouse scale, full observability stack, load testing, DR drill",                    "DevOps + Platform Engineer",        "Zero-downtime DR drill · P95 agent latency < 30s · Load test 100 users"),
]

phase_colors_sp = {
    "Ph 1": ACCENT_BLUE, "Ph 2": ACCENT_TEAL, "Ph 3": ACCENT_PURP,
    "Ph 4": ACCENT_ORG,  "Ph 5": ACCENT_GOLD, "Ph 6": "059669", "Ph 7": DARK_NAVY,
}

for row_i, spr in enumerate(SPRINTS, start=3):
    phase_key = spr[2][:4]
    bg = phase_colors_sp.get(phase_key, ACCENT_BLUE)
    for col_i, val in enumerate(spr, start=1):
        cell = ws4.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        if col_i <= 3:
            cell.fill = hfill(bg)
            cell.font = bold(9, WHITE)
            cell.alignment = wrap_center()
        else:
            cell.fill = hfill(WHITE if row_i % 2 else "F5F3FF")
            cell.font = reg(9)
            cell.alignment = wrap_left()
    ws4.row_dimensions[row_i].height = 42

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – TECH STACK
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⚙️ Tech Stack")
ws5.sheet_view.showGridLines = False

ts_cols = ["Layer", "Technology", "Version", "Role in System", "Why Chosen", "Alternatives Considered", "Self-Hosted?"]
ts_widths = [20, 22, 14, 38, 36, 28, 14]
for i, w in enumerate(ts_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells(f"A1:{get_column_letter(len(ts_cols))}1")
ws5["A1"] = "COMPLETE TECHNOLOGY STACK — VERSIONS, RATIONALE & ALTERNATIVES"
ws5["A1"].fill = hfill(DARK_NAVY)
ws5["A1"].font = bold(13, WHITE)
ws5["A1"].alignment = wrap_center()
ws5.row_dimensions[1].height = 26

for i, h in enumerate(ts_cols, 1):
    c = ws5.cell(row=2, column=i, value=h)
    c.fill = hfill(ACCENT_ORG)
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws5.row_dimensions[2].height = 26

TECH_STACK = [
    ("🤖 AI/LLM","LangGraph","0.2+","Stateful multi-agent orchestration; defines agent state machines, tool calling, and inter-agent flow","Best-in-class stateful LLM orchestration with first-class Python support; native LangChain integration","AutoGen / CrewAI / custom orchestration","Yes (library)"),
    ("🤖 AI/LLM","LiteLLM","Latest","Multi-provider LLM gateway: routes calls to OpenAI/Anthropic/Gemini with fallback, cost tracking, and caching","Single interface to 100+ LLM providers; transparent cost tracking; built-in fallback chains","Direct SDK calls (no fallback)","Yes (proxy server)"),
    ("🤖 AI/LLM","DSPy","2.x","Structured LLM outputs for agents; prompt optimization via bootstrapping; confidence calibration","Replaces fragile f-string prompts with compiled, optimized modules; improves consistency 30-40%","Guidance / Instructor / raw prompting","Yes (library)"),
    ("🔧 Backend","FastAPI","0.111+","Primary API framework for all microservices; async-first, OpenAPI auto-generation, dependency injection","Fastest Python framework for async APIs; native Pydantic validation; excellent for microservices","Django REST / Flask / Litestar","Yes"),
    ("🔧 Backend","Python","3.11+","Primary application language for all backend services and AI agents","Best ecosystem for AI/ML; excellent async support; team familiarity","Go (for performance-critical services if needed)","Yes"),
    ("🔧 Backend","Celery","5.3+","Async task queue for background jobs (email send, social publishing, data sync, file processing)","Battle-tested; Redis/RabbitMQ broker support; excellent monitoring (Flower)","Dramatiq / RQ (Redis Queue)","Yes"),
    ("🎨 Frontend","Next.js","14 (App Router)","Full-stack React framework for the web application; SSR for SEO, App Router for layouts","Industry standard for React production apps; excellent DX; Vercel deployment support","Remix / Nuxt.js / SvelteKit","Yes (self-hosted)"),
    ("🎨 Frontend","React","18.x","Core UI component library; Suspense and concurrent features for smooth UX","Dominant ecosystem; massive component library availability; team expertise","Vue 3 / Svelte","Yes (via Next.js)"),
    ("🎨 Frontend","Tailwind CSS","3.4+","Utility-first CSS framework for all UI styling","Rapid development; consistent design system; eliminates CSS specificity issues","Styled Components / MUI / Chakra UI","Yes (library)"),
    ("🎨 Frontend","React Flow","11+","Visual drag-and-drop workflow builder canvas","Best-in-class graph/flow visualization for React; extensive customization; MIT license","Reactflow alternatives / D3.js custom","Yes (library)"),
    ("🎨 Frontend","Recharts","2.x","Data visualization charts for analytics dashboards","Well-integrated with React; responsive; composable chart components","Chart.js / D3.js / Victory Charts","Yes (library)"),
    ("🗃 Database","PostgreSQL","15+","Primary relational database for all structured application data; Row-Level Security for multi-tenancy","ACID compliance; RLS for tenant isolation at DB layer; excellent JSON support; battle-tested","MySQL / CockroachDB","Yes"),
    ("🗃 Database","Qdrant","1.9+","Primary vector database for semantic search, RAG retrieval, and AI memory","Best performance-to-cost ratio; tenant-isolated collections; excellent filtering; Rust-based","Pinecone (expensive) / Weaviate / Chroma","Yes"),
    ("🗃 Database","Redis","7+","Session store, rate limiting counters, Celery broker, short-term agent memory cache, distributed locks","Sub-millisecond latency; versatile data structures; pub/sub for real-time events","Memcached (less features) / Dragonfly","Yes"),
    ("🗃 Database","ClickHouse","24.x","Time-series analytics database for campaign performance metrics and high-volume event data","Columnar storage for 100x faster analytics queries vs. Postgres; excellent compression","TimescaleDB / Apache Druid / BigQuery","Yes"),
    ("⚙️ Workflow","Temporal","1.22+","Durable workflow execution engine for complex marketing automation sequences","Guaranteed execution with automatic retries; native human-in-the-loop signaling; battle-tested at Uber","Apache Airflow (not suited for HITL) / Prefect","Yes"),
    ("📨 Messaging","Apache Kafka","3.6+","Event streaming for audit logs, inter-agent communication, analytics event pipeline, async workflows","High-throughput, durable, replay-capable event streaming; backbone for async architecture","RabbitMQ (lower throughput) / AWS SQS","Yes"),
    ("📨 Messaging","RabbitMQ / Celery","3.12+ / 5.3+","Task queue for discrete async jobs (publishing, notifications, data sync)","Simpler than Kafka for task queues; Celery provides worker management and retry logic","Celery with Redis broker only","Yes"),
    ("🔒 Security","Open Policy Agent","0.65+","Policy-as-code engine for the CGH governance layer; evaluates all agent actions against policy rules","Declarative Rego policies; decouples policy from application code; auditable","Custom middleware / Casbin","Yes"),
    ("🔒 Security","HashiCorp Vault","1.15+","Secrets management: stores API keys, DB credentials, OAuth tokens, encryption keys","Industry standard; dynamic secrets; automatic rotation; K8s integration via AppRole","AWS Secrets Manager / Azure Key Vault","Yes (OSS)"),
    ("🔒 Security","Istio","1.21+","Service mesh providing mTLS between all services, traffic management, observability","Zero-trust networking; mTLS without app code changes; circuit breaking; telemetry","Linkerd (simpler) / Consul Connect","Yes"),
    ("☁️ Infra","Kubernetes","1.29+","Container orchestration for all microservices; auto-scaling, self-healing, rolling deployments","Industry standard; excellent ecosystem; cloud-agnostic; Helm for package management","Docker Swarm (less scalable) / Nomad","Yes (via cloud)"),
    ("☁️ Infra","Terraform","1.8+","Infrastructure-as-code for all cloud resources (K8s, DBs, networking, DNS)","Declarative, version-controlled infra; state management; provider ecosystem","Pulumi / CDK / Ansible","Yes (tool)"),
    ("☁️ Infra","Docker","25+","Container packaging for all services; multi-stage builds for optimized images","Universal container standard; multi-stage builds reduce image size by 70%+","Podman","Yes"),
    ("📊 Observability","Grafana + Prometheus","10.x / 2.x","Metrics collection and visualization; 200+ custom metrics across all services","Open-source; excellent K8s integration; large dashboard library; mature alerting","Datadog (expensive) / New Relic","Yes"),
    ("📊 Observability","Jaeger / Tempo","1.57 / 2.x","Distributed tracing for request flows across microservices","Understand end-to-end latency; trace agent calls across services; identify bottlenecks","Zipkin / AWS X-Ray","Yes"),
    ("📊 Observability","MLflow","2.x","ML experiment tracking for A/B prompt tests, model versions, performance metrics","Open-source; simple Python API; excellent UI; artifact storage","Weights & Biases (cost) / Neptune","Yes"),
    ("🗂 DevOps","GitHub Actions","Latest","CI/CD pipeline: test, build, scan, deploy on every PR and push to main","Deep GitHub integration; generous free tier; marketplace actions","GitLab CI / Jenkins / CircleCI","Yes (GitHub-hosted)"),
    ("🗂 DevOps","Helm","3.x","Kubernetes package manager for deploying and versioning all service configurations","Templating for K8s manifests; rollback support; chart versioning","Kustomize / raw manifests","Yes (tool)"),
]

layer_colors = {
    "🤖 AI/LLM": ACCENT_PURP, "🔧 Backend": ACCENT_BLUE, "🎨 Frontend": ACCENT_TEAL,
    "🗃 Database": ACCENT_ORG, "⚙️ Workflow": ACCENT_GOLD, "📨 Messaging": "1C64F2",
    "🔒 Security": "991B1B", "☁️ Infra": "374151", "📊 Observability": "065F46",
    "🗂 DevOps": "4B5563",
}

for row_i, tech in enumerate(TECH_STACK, start=3):
    layer_key = tech[0]
    bg = layer_colors.get(layer_key, ACCENT_BLUE)
    for col_i, val in enumerate(tech, start=1):
        cell = ws5.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        if col_i == 1:
            cell.fill = hfill(bg)
            cell.font = bold(8, WHITE)
            cell.alignment = wrap_center()
        elif col_i == 2:
            cell.fill = hfill(bg)
            cell.font = bold(9, WHITE)
            cell.alignment = wrap_left()
        else:
            cell.fill = hfill(WHITE if row_i % 2 else "FFF7ED")
            cell.font = reg(9)
            cell.alignment = wrap_left()
    ws5.row_dimensions[row_i].height = 44

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – RISK REGISTER
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("⚠️ Risk Register")
ws6.sheet_view.showGridLines = False

rk_cols = ["Risk ID", "Category", "Risk Description", "Probability", "Impact", "Risk Score",
           "Phase Affected", "Mitigation Strategy", "Contingency Plan", "Owner", "Status"]
rk_widths = [10, 18, 40, 12, 12, 12, 16, 42, 36, 20, 12]
for i, w in enumerate(rk_widths, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.merge_cells(f"A1:{get_column_letter(len(rk_cols))}1")
ws6["A1"] = "COMPREHENSIVE RISK REGISTER — PROBABILITY × IMPACT MATRIX"
ws6["A1"].fill = hfill(DARK_NAVY)
ws6["A1"].font = bold(13, WHITE)
ws6["A1"].alignment = wrap_center()
ws6.row_dimensions[1].height = 26

for i, h in enumerate(rk_cols, 1):
    c = ws6.cell(row=2, column=i, value=h)
    c.fill = hfill("991B1B")
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws6.row_dimensions[2].height = 26

RISKS = [
    ("R-001","AI / LLM","LLM hallucinations generating inaccurate marketing strategies or brand-damaging content","High","High","🔴 Critical","Phase 3–6","DSPy structured outputs with schema validation · Confidence scoring thresholds · Human-in-the-loop approval gate mandatory for all AI outputs before publishing","Fallback to template-based content generation if confidence < threshold · Automatic flagging for human review","AI Lead","Active"),
    ("R-002","AI / Security","Prompt injection attacks via user-supplied content manipulating agent behavior","Medium","High","🟠 High","Phase 3–7","CGH semantic prompt guards on all user inputs · Input sanitization layer · OPA policy validation of agent tool calls · Separate input context from system context","Quarantine suspicious inputs · Alert security team · Block affected agent temporarily","Security Lead","Active"),
    ("R-003","Data Security","Cross-tenant data leakage via RLS bypass or vector DB misconfiguration","Low","Critical","🟠 High","Phase 1–7","PostgreSQL RLS on all tables · Tenant-isolated Qdrant collections enforced at API layer · Automated cross-tenant penetration testing in CI","Immediate tenant data isolation · Incident response plan · Regulatory notification if GDPR breach","Security Lead","Active"),
    ("R-004","Infrastructure","Temporal workflow state corruption causing campaign automation failures","Low","High","🟡 Medium","Phase 4–7","Temporal's built-in event sourcing and durable execution · Regular Temporal DB backups · Crash recovery load tests","Replay workflows from last checkpoint · Manual intervention UI for stuck workflows","DevOps Lead","Active"),
    ("R-005","Integration","Social media API breaking changes (Meta, LinkedIn, X) disrupting publishing","High","Medium","🟠 High","Phase 5–7","Version-pinned API clients · API deprecation monitoring service · Modular client architecture · Webhook on API change alerts","Immediate switch to manual publishing fallback · Buffer API as backup publisher","Backend Lead","Active"),
    ("R-006","Performance","Vector search latency degradation under high concurrent load","Medium","Medium","🟡 Medium","Phase 2–7","Qdrant HNSW index tuning · Collection sharding at 1M vectors · Distributed Qdrant cluster · Query result caching in Redis","Horizontal scaling of Qdrant pods · Temporary search quality reduction for speed","Platform Lead","Active"),
    ("R-007","Compliance","GDPR data subject access request volume overwhelming manual process","Medium","High","🟠 High","Phase 7+","Automated DSAR pipeline (30-day SLA target) · Data inventory automation · Pre-built erasure workflows","Outsource DSAR processing temporarily · Engage DPA for extension in exceptional cases","Legal/Backend Lead","Active"),
    ("R-008","Cost","LLM API costs exceeding budget projections with scale","High","Medium","🟠 High","Phase 3–7","LiteLLM cost tracking per tenant · Token budget enforcement per request · Prompt optimization reducing avg tokens 40% · Caching identical requests","Switch to smaller/cheaper models for lower-priority tasks · Implement per-tenant token quotas","Product Lead","Active"),
    ("R-009","Technical","Over-engineered microservice architecture increasing operational complexity","Medium","Medium","🟡 Medium","Phase 1–4","Start with modular monolith per domain; split only when needed · Comprehensive runbooks · On-call rotation · Observability-first design","Consolidate underutilized services · Hire additional DevOps if operational overhead exceeds threshold","Tech Lead","Active"),
    ("R-010","Integration","Third-party API rate limiting blocking automated publishing during peak times","High","High","🔴 Critical","Phase 5–7","Per-platform token bucket rate limiting · Celery queue with exponential backoff · Intelligent scheduling to spread requests","Priority queue: paid ads over organic posts · Delay non-critical operations · Notify users of delays","Backend Lead","Active"),
    ("R-011","AI / ML","Agent confidence calibration degrading over time (model drift)","Medium","Medium","🟡 Medium","Phase 6+","Weekly automated calibration evaluation jobs · Drift detection alerts at ECE > 0.1 · Monthly human evaluation sample","Trigger DSPy re-optimization · Roll back to last stable prompt version if quality drops","AI Lead","Active"),
    ("R-012","Business","Key third-party services (Stripe, Twilio) having significant outages","Low","High","🟡 Medium","Phase 5+","Multi-provider architecture (Stripe primary / Paddle fallback for payments) · Circuit breakers · Graceful degradation (cache last successful response)","Activate fallback payment processor · Communicate to users · SLA compensation","Backend Lead","Active"),
]

risk_colors = {"🔴 Critical": "FEE2E2", "🟠 High": "FEF3C7", "🟡 Medium": "FEFCE8", "🟢 Low": "ECFDF5"}

for row_i, risk in enumerate(RISKS, start=3):
    score = risk[5]
    row_bg = risk_colors.get(score, WHITE)
    for col_i, val in enumerate(risk, start=1):
        cell = ws6.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        cell.fill = hfill(row_bg)
        cell.font = bold(9) if col_i in (1, 5) else reg(9)
        cell.alignment = wrap_center() if col_i in (1,4,5,6,7,10,11) else wrap_left()
    ws6.row_dimensions[row_i].height = 56

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – DEPENDENCY MAP
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("🔗 Dependencies")
ws7.sheet_view.showGridLines = False

dep_cols = ["Task ID", "Task Name", "Depends On (Task IDs)", "Downstream Dependents", "Blocking?", "Notes"]
dep_widths = [10, 36, 28, 36, 12, 42]
for i, w in enumerate(dep_widths, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

ws7.merge_cells(f"A1:{get_column_letter(len(dep_cols))}1")
ws7["A1"] = "TASK DEPENDENCY MAP — CRITICAL PATH & BLOCKING RELATIONSHIPS"
ws7["A1"].fill = hfill(DARK_NAVY)
ws7["A1"].font = bold(13, WHITE)
ws7["A1"].alignment = wrap_center()
ws7.row_dimensions[1].height = 26

for i, h in enumerate(dep_cols, 1):
    c = ws7.cell(row=2, column=i, value=h)
    c.fill = hfill("065F46")
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws7.row_dimensions[2].height = 26

DEPS = [
    ("1.1","Auth & Identity Service","None (Phase 1 start)","1.2, 1.3, all authenticated endpoints","YES","Foundation of entire system — must be completed first"),
    ("1.2","Multi-Tenant Management & RLS","1.1 (User identity required for tenant association)","2.1, 2.2, 2.3, 3.1, all data services","YES","RLS policies block all data services from starting"),
    ("1.3","API Gateway & Service Mesh","1.1 (Auth middleware needed for gateway)","All microservices (external access)","YES","All services route through gateway"),
    ("1.4","CI/CD Pipeline","1.1, 1.2, 1.3 (services to build and deploy)","All subsequent phases (deployment pipeline)","NO","Can be built in parallel; blocks production deployments"),
    ("2.1","Document Ingestion Pipeline","1.2 (tenant isolation), 1.3 (gateway)","2.2, 2.3 (vectors stored)","YES","No ingestion = no vectors = no RAG"),
    ("2.2","Vector DB & Semantic Search","2.1 (vectors to store and search)","2.3, 3.1, 3.2 (all RAG-dependent agents)","YES","All RAG agents depend on semantic search"),
    ("2.3","RAG Pipeline & Memory Service","2.1, 2.2 (retrieval infrastructure)","3.1, 3.2, 3.3, 3.4, 3.5 (all agents use RAG)","YES","RAG pipeline is the intelligence layer for all agents"),
    ("2.4","Brand Knowledge Base","2.3 (embedding pipeline)","3.2, 3.3, 3.4 (brand context injection)","YES","Without brand context, agent outputs will be generic"),
    ("3.1","Agent Orchestration Framework","2.3, 2.4 (memory and RAG ready)","3.2, 3.3, 3.4, 3.5 (all agents)","YES","All agents register in this framework"),
    ("3.2","Marketing Strategy Agent","3.1 (orchestration framework)","3.3, 3.4 (content/SEO use strategies)","YES","Strategy outputs drive all content and SEO agents"),
    ("3.3","Content Ideas Generator Agent","3.1, 3.2 (strategy context required)","4.2 (scheduled content), 6.1 (analytics)","NO","Can run independently once orchestration is ready"),
    ("3.4","SEO Strategy Agent","3.1, 2.3 (keyword research via RAG)","6.1 (SEO performance tracking)","NO","Can be built in parallel with 3.3"),
    ("3.5","7 Additional Agents","3.1 (registry and base class)","4.2, 5.3, 5.4, 6.1, 6.2","NO","Can be built in parallel after orchestration framework"),
    ("4.1","Temporal Infrastructure","3.1 (agents to orchestrate)","4.2, 4.3, 4.4 (all automation)","YES","Entire automation layer depends on Temporal"),
    ("4.2","Campaign Scheduling Engine","4.1, 5.2 (platform connections needed for publishing)","6.1 (performance tracking of published content)","YES","Publishing requires platform API connections from Phase 5"),
    ("4.3","Visual Workflow Builder","4.1 (Temporal to trigger workflows)","4.4 (HITL nodes in workflow builder)","NO","UI enhancement; can be built in parallel with 4.2"),
    ("4.4","HITL Approval System","4.1 (Temporal signals), 4.3 (UI)","All published outputs (gate before publish)","YES","No publishing should happen without HITL gate"),
    ("5.1","Integration Hub & Vault","1.3 (gateway for routing), 4.1 (workflow integration)","5.2, 5.3, 5.4, 5.5, 5.6","YES","All integrations use Vault for credential storage"),
    ("5.2","Social Media Integrations","5.1 (Vault for credentials)","4.2 (publishing requires these connections), 6.1","YES","Campaign scheduling requires live platform connections"),
    ("5.3","Advertising Integrations","5.1 (Vault), 3.5 (Ads agent)","6.1 (ads performance analytics)","NO","Ads integrations are parallel to social media"),
    ("5.4","Analytics & CRM Integrations","5.1 (Vault)","6.1 (data for analytics dashboard)","NO","Analytics integrations are parallel to ads"),
    ("5.5","Email Integrations","5.1 (Vault), 3.5 (Email agent)","6.1 (email performance tracking)","NO","Parallel to other integrations"),
    ("5.6","Stripe Payments","1.2 (tenant management), 5.1 (Vault)","All features gated by subscription tier","YES","Feature access requires subscription status"),
    ("6.1","Campaign Analytics Dashboard","5.2, 5.3, 5.4 (data sources), 4.2 (published posts)","6.2, 6.3 (optimization uses analytics data)","NO","Analytics is value-add layer on top of integrations"),
    ("6.2","AI Feedback Loops","3.1, 6.1 (performance data for feedback)","6.3 (A/B tests use optimized prompts)","NO","Optimization layer requires performance data"),
    ("6.3","A/B Prompt Testing","6.2 (feedback infrastructure)","3.x agents (improved prompts)","NO","Enhancement; can be built after core agents work"),
    ("6.4","Dashboard UI","3.2, 4.2, 5.1–5.5, 6.1 (all data APIs)","User-facing access to all features","NO","UI is built on top of completed backend APIs"),
    ("6.5","Content Calendar UI","4.2 (scheduling backend), 3.3 (content ideas)","User publishing workflow","NO","UI layer on scheduling backend"),
    ("6.6","Notifications & Settings UI","All backend services (to configure)","User experience completeness","NO","Can be built in parallel with other UI work"),
    ("7.1","CGH Full Implementation","1.3 (Istio mesh), 3.1 (agent framework)","7.2, 7.3 (governance layer)","YES","CGH is the compliance foundation"),
    ("7.2","SSO & Audit Signing","7.1 (CGH audit pipeline), 5.6 (billing for enterprise)","Enterprise client onboarding","NO","Enterprise-tier feature"),
    ("7.3","GDPR & SOC2 Tooling","7.1 (CGH), 5.4 (data sources for compliance)","Commercial operations (customer trust)","NO","Required for enterprise sales"),
    ("7.4","Multi-Region K8s","7.1 (full compliance), 1.4 (CI/CD)","Production global deployment","NO","Scale initiative; can run in parallel"),
    ("7.5","Kafka + ClickHouse Scale","7.4 (infra foundation), 6.1 (analytics needs scale)","High-volume analytics at scale","NO","Upgrade path for existing Kafka/ClickHouse from earlier phases"),
    ("7.6","Observability Stack","All services (to instrument)","Operations team effectiveness","NO","Can be instrumented incrementally throughout all phases"),
]

for row_i, dep in enumerate(DEPS, start=3):
    is_blocking = dep[4] == "YES"
    for col_i, val in enumerate(dep, start=1):
        cell = ws7.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        if col_i == 5:
            cell.fill = hfill("FEE2E2") if is_blocking else hfill("ECFDF5")
            cell.font = bold(9, "991B1B") if is_blocking else bold(9, "065F46")
            cell.alignment = wrap_center()
        else:
            cell.fill = hfill(WHITE if row_i % 2 else "F0FDF4")
            cell.font = bold(9) if col_i == 2 else reg(9)
            cell.alignment = wrap_left()
    ws7.row_dimensions[row_i].height = 38

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 – TEAM & RESOURCES
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("👥 Team & Resources")
ws8.sheet_view.showGridLines = False

for col, w in zip("ABCDEFG", [22, 30, 14, 22, 32, 18, 22]):
    ws8.column_dimensions[col].width = w

ws8.merge_cells("A1:G1")
ws8["A1"] = "TEAM STRUCTURE, ROLES & RESOURCE PLANNING"
ws8["A1"].fill = hfill(DARK_NAVY)
ws8["A1"].font = bold(13, WHITE)
ws8["A1"].alignment = wrap_center()
ws8.row_dimensions[1].height = 26

team_headers = ["Role", "Responsibilities", "Headcount", "Key Skills Required", "Primary Phases", "Tool Access", "Notes"]
for i, h in enumerate(team_headers, 1):
    c = ws8.cell(row=2, column=i, value=h)
    c.fill = hfill(ACCENT_GOLD)
    c.font = bold(9, WHITE)
    c.alignment = wrap_center()
    c.border = thin_border()
ws8.row_dimensions[2].height = 26

TEAM = [
    ("Senior AI Engineer","Agent orchestration (LangGraph), RAG pipeline, DSPy optimization, agent state machines, multi-agent coordination","2","Python 3.11, LangGraph, LiteLLM, DSPy, LangChain, RAG architecture, vector databases","Phase 2, 3, 6","Full codebase + Qdrant admin + LiteLLM config","Critical hire; scarce talent; hire 6+ months before Phase 3 start"),
    ("Senior Backend Engineer","Core auth service, tenant management, API design, database schema, microservice architecture","2","FastAPI, PostgreSQL, Redis, Alembic, Celery, microservices, SQLAlchemy","Phase 1, 2, 4, 5","All backend services + DB admin","Lead developer for Phase 1 critical path"),
    ("Frontend Engineer","Next.js app, React components, dashboard UI, content calendar, settings pages, responsive design","2","Next.js 14, React 18, TypeScript, Tailwind CSS, React Flow, Recharts, Zustand","Phase 6 (primarily) + UI in all phases","Frontend codebase + Figma + Vercel","1 senior (workflow builder), 1 mid-level (standard UI pages)"),
    ("DevOps / Platform Engineer","K8s infrastructure, CI/CD, Docker, Helm charts, multi-region deployment, Terraform, observability stack","2","Kubernetes, Terraform, Docker, Helm, Istio, GitHub Actions, Grafana, Prometheus","Phase 1, 7 (primarily) + ongoing all phases","Full infra + Vault + K8s cluster admin","1 senior for Phase 1 infra; 2nd engineer from Phase 7"),
    ("Security Engineer","OPA policies (CGH), Vault setup, mTLS config, SSO/SAML, audit signing, penetration testing","1","OPA/Rego, HashiCorp Vault, SAML 2.0, OIDC, K8s security, threat modeling, GDPR","Phase 1 (auth), Phase 7 (governance)","All security systems + audit logs + Vault admin","Can be part-time Phase 1–6; full-time Phase 7"),
    ("ML/AI Engineer","RAG pipeline tuning, embedding model evaluation, confidence calibration, MLflow tracking, A/B prompt experiments","1","Python, RAG, RAGAS evaluation, MLflow, DSPy, pandas, LLM evaluation frameworks","Phase 2, 5, 6","AI/ML stack + MLflow + Qdrant read","Can overlap with Senior AI Engineer role in smaller teams"),
    ("QA / Test Engineer","Automated test suite (pytest), integration tests, load testing (k6), security testing, cross-tenant penetration tests","1","pytest, k6, Postman, automated testing, security testing basics, CI integration","All phases (QA starts Phase 1)","All test environments + CI pipeline","Start QA from day 1; not an afterthought"),
    ("Product Manager","Requirements management, sprint planning, stakeholder communication, user story writing, roadmap","1","Agile/Scrum, product thinking, technical communication, JIRA/Linear","All phases","JIRA/Linear + all documentation","Technical PM strongly preferred for AI product"),
    ("Technical Project Manager","Delivery tracking, dependency management, risk management, resource planning, executive reporting","1","Project management, risk management, technical understanding, communication","All phases","JIRA/Linear + all dashboards","Overlap with PM role in smaller teams"),
]

for row_i, member in enumerate(TEAM, start=3):
    for col_i, val in enumerate(member, start=1):
        cell = ws8.cell(row=row_i, column=col_i, value=val)
        cell.border = thin_border()
        cell.fill = hfill(WHITE if row_i % 2 else "FFFBEB")
        cell.font = bold(9) if col_i == 1 else reg(9)
        cell.alignment = wrap_left()
    ws8.row_dimensions[row_i].height = 58

# ── Change log ────────────────────────────────────────────────────────────
ws_log = wb.create_sheet("📝 Change Log")
ws_log.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [12, 20, 26, 50, 20]):
    ws_log.column_dimensions[col].width = w
ws_log.merge_cells("A1:E1")
ws_log["A1"] = "VERSION HISTORY & CHANGE LOG"
ws_log["A1"].fill = hfill(DARK_NAVY)
ws_log["A1"].font = bold(12, WHITE)
ws_log["A1"].alignment = wrap_center()
ws_log.row_dimensions[1].height = 24

for i, h in enumerate(["Version","Date","Author","Changes Summary","Status"], 1):
    c = ws_log.cell(row=2, column=i, value=h)
    c.fill = hfill(ACCENT_BLUE); c.font = bold(9, WHITE); c.alignment = wrap_center(); c.border = thin_border()

logs = [
    ("1.0","2024-05-20","Senior Technical PM","Initial enterprise roadmap release — basic phase structure","Superseded"),
    ("2.0","June 2026","Senior Technical PM","Complete re-architecture: 35 tasks across 7 phases, full 3rd party integration registry (40 services), sprint plan, risk register, dependency map, team structure, detailed security controls per task","Current"),
]
for row_i, log in enumerate(logs, start=3):
    for col_i, val in enumerate(log, start=1):
        cell = ws_log.cell(row=row_i, column=col_i, value=val)
        cell.fill = hfill(WHITE if row_i % 2 else LIGHT_GRAY)
        cell.font = reg(9); cell.alignment = wrap_left(); cell.border = thin_border()
    ws_log.row_dimensions[row_i].height = 44

# ── Save ───────────────────────────────────────────────────────────────────
out = "AI_Marketing_OS_Roadmap_v2.xlsx"
wb.save(out)
print(f"Saved -> {out}")
