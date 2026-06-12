import json
import logging
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.future import select
from shared.database import async_session_maker
from shared.models.tenant import CompanyContext, Tenant, StrategyPlan

from .base_agent import BaseMarketingAgent, AgentState
from ..registry import AgentManifest
from shared.utils.ai_integration import AIClient

logger = logging.getLogger(__name__)

MANIFEST = AgentManifest(
    agent_id="strategy_agent",
    name="Marketing Strategy Planning Agent",
    version="2.0.0",
    description="Creates complete execution strategy plans and exports them into structured JSON and Excel reports.",
    capabilities=[
        "prompt_analysis",
        "company_analysis",
        "strategy_planning",
        "execution_planning",
        "goal_mapping",
        "timeline_creation",
        "excel_export"
    ],
    tools_required=["web_search"],
    tenant_plan_required="starter"
)

# ─────────────────────────────────────────────────────────────
# Console helpers
# ─────────────────────────────────────────────────────────────
def _banner(title: str):
    bar = "─" * 60
    print(f"\n{bar}")
    print(f"  🔷 {title}")
    print(bar)

def _step(label: str, value=None):
    print(f"  ▶ {label}")
    if value is not None:
        if isinstance(value, (dict, list)):
            print(json.dumps(value, indent=4, ensure_ascii=False))
        else:
            print(f"    {value}")

def _ok(label: str, value=None):
    print(f"  ✅ {label}")
    if value is not None:
        if isinstance(value, (dict, list)):
            print(json.dumps(value, indent=4, ensure_ascii=False))
        else:
            print(f"    {value}")

def _err(label: str, error=None):
    print(f"  ❌ {label}")
    if error:
        print(f"    ERROR: {error}")


class StrategyAgent(BaseMarketingAgent):

    SYSTEM_PROMPT = """
    You are an elite AI Marketing Strategy Planner.
    You will receive a specific sub-task to generate for the overall marketing strategy.
    Generate ONLY valid JSON. Do not include markdown formatting like ```json.
    Ensure your response exactly matches the requested structure.
    """

    # ─────────────────────────────────────────────────────────────
    # STEP 1: PLAN
    # ─────────────────────────────────────────────────────────────
    async def _plan(self, state: AgentState) -> AgentState:
        task = state["task"]

        if "extra_data" not in state or not state["extra_data"]:
            _banner("STEP 1 — AGENT INITIALIZING")
            _step("User prompt", task.description or task.data.get("description", ""))
            _step("Tenant ID", state["tenant_id"])

            product_id = task.data.get("product_id")

            state["extra_data"] = {
                "product_id": product_id,
                "strategy_plan": {
                    "company_name": "Unknown",
                    "industry": "Unknown",
                    "target_audience": "Unknown",
                    "created_at": datetime.utcnow().isoformat(),
                    "strategy": {
                        "business_goal": "",
                        "marketing_goal": "",
                        "core_strategy": []
                    },
                    "phases": [],
                    "channels": [],
                    "execution_plan": []
                },
                "current_task": ""
            }

            state["plan"] = [
                "basic_info",
                "core_strategy",
                "phases",
                "channels",
                "execution_plan"
            ]

            user_prompt = task.description or task.data.get("description", "")
            state["extra_data"]["user_prompt"] = user_prompt

            _ok("Task queue initialized", state["plan"])
            logger.info("Initialized task list for StrategyAgent")

        # Pop next task
        if state.get("plan"):
            next_task = state["plan"].pop(0)
            state["extra_data"]["current_task"] = next_task
            remaining = state.get("plan", [])
            _banner(f"PLAN — Next Sub-Task: '{next_task.upper()}'")
            _step("Remaining queue", remaining if remaining else "(empty — this is the last task)")
            logger.info(f"Popped next task: {next_task}")

        return state

    # ─────────────────────────────────────────────────────────────
    # STEP 2: EXECUTE
    # ─────────────────────────────────────────────────────────────
    async def _execute(self, state: AgentState) -> AgentState:
        current_task = state["extra_data"].get("current_task")
        user_prompt = state["extra_data"].get("user_prompt", "")
        tenant_id = state["tenant_id"]
        iteration = state.get("iteration", 0) + 1

        _banner(f"EXECUTE [{iteration}] — Sub-Task: '{current_task.upper()}'")
        logger.info(f"Executing task: {current_task}")

        # ── BASIC INFO: DB fetch, no LLM ──────────────────────────
        if current_task == "basic_info":
            _step("Source", "Fetching company data directly from database (no LLM)")
            try:
                product_id = state["extra_data"].get("product_id")
                product_context = ""
                
                async with async_session_maker() as session:
                    tenant_result = await session.execute(
                        select(Tenant).where(Tenant.id == tenant_id)
                    )
                    tenant_record = tenant_result.scalars().first()

                    ctx_result = await session.execute(
                        select(CompanyContext).where(CompanyContext.tenant_id == tenant_id)
                    )
                    ctx = ctx_result.scalars().first()

                    if product_id:
                        from shared.models.tenant import CompanyProduct
                        prod_result = await session.execute(
                            select(CompanyProduct).where(CompanyProduct.id == product_id)
                        )
                        prod_record = prod_result.scalars().first()
                        if prod_record:
                            product_context = (
                                f"--- TARGET PRODUCT/SERVICE DETAILS ---\n"
                                f"Name: {prod_record.name}\n"
                                f"Type: {prod_record.type}\n"
                                f"Description: {prod_record.description or ''}\n"
                                f"Target Audience: {prod_record.target_audience or ''}\n"
                                f"Features: {prod_record.features or ''}\n"
                            )
                            state["extra_data"]["product_name"] = prod_record.name
                            state["extra_data"]["product_type"] = prod_record.type
                            state["extra_data"]["product_description"] = prod_record.description
                            state["extra_data"]["product_target_audience"] = prod_record.target_audience
                            state["extra_data"]["product_features"] = prod_record.features

                company_name = (tenant_record.name if tenant_record else None) or "Unknown"
                industry     = (ctx.focus if ctx else None) or "General"
                
                if product_id and "product_name" in state["extra_data"]:
                    target_audience = state["extra_data"].get("product_target_audience") or "General Audience"
                else:
                    target_audience = (
                        (ctx.product_details if ctx else None)
                        or (ctx.service_details if ctx else None)
                        or "General Audience"
                    )

                state["extra_data"]["strategy_plan"]["company_name"]    = company_name
                state["extra_data"]["strategy_plan"]["industry"]        = industry
                state["extra_data"]["strategy_plan"]["target_audience"] = target_audience
                state["extra_data"]["product_context"] = product_context

                _ok("Company data loaded from DB", {
                    "company_name":    company_name,
                    "industry":        industry,
                    "target_audience": target_audience,
                    "product_id":      product_id
                })
                logger.info(
                    f"Loaded company data: name='{company_name}', "
                    f"industry='{industry}', tenant={tenant_id}"
                )
            except Exception as e:
                _err("Failed to fetch company data from DB", e)
                logger.error(f"Failed to fetch company data from DB: {e}")

            state["iteration"] = iteration
            state["confidence"] = 0.0
            return state

        # ── LLM TASKS ─────────────────────────────────────────────
        prompt = ""
        product_context = state["extra_data"].get("product_context", "")
        if current_task == "core_strategy":
            prompt = f"""
            Analyze the user prompt: '{user_prompt}'.
            {product_context}
            Define the core marketing strategy.
            Return a JSON object with EXACTLY these keys:
            {{
                "business_goal": "The overarching business objective",
                "marketing_goal": "The specific marketing objective",
                "core_strategy": ["Strategy point 1", "Strategy point 2", "Strategy point 3"]
            }}
            """
        elif current_task == "phases":
            prompt = f"""
            Based on the strategy for '{user_prompt}', define 4 execution phases.
            {product_context}
            Return a JSON array of objects, where each object has EXACTLY these keys:
            [
                {{
                    "phase": "Phase Name (e.g., Phase 1 - Setup)",
                    "duration": "Duration (e.g., Week 1-2)",
                    "tasks": ["Task 1", "Task 2"]
                }}
            ]
            """
        elif current_task == "channels":
            prompt = f"""
            Based on the strategy for '{user_prompt}', define 4 marketing channels.
            {product_context}
            Return a JSON array of objects, where each object has EXACTLY these keys:
            [
                {{
                    "channel": "Channel Name (e.g., SEO, Google Ads)",
                    "objective": "Channel specific objective",
                    "kpi": "Key Performance Indicator"
                }}
            ]
            """
        elif current_task == "execution_plan":
            prompt = f"""
            Based on the strategy for '{user_prompt}', create a 4-week execution timeline.
            {product_context}
            Return a JSON array of objects, where each object has EXACTLY these keys:
            [
                {{
                    "week": "Week Number (e.g., Week 1)",
                    "activity": "Main activity",
                    "owner": "Who owns this (e.g., Marketing Team)",
                    "status": "Pending"
                }}
            ]
            """

        if prompt:
            _step("Sending prompt to AI (OpenRouter)...")
            try:
                raw_result = await AIClient.generate_completion(tenant_id, prompt, self.SYSTEM_PROMPT)

                print("\n  📨 RAW AI RESPONSE:")
                print("  " + "·" * 50)
                print(raw_result)
                print("  " + "·" * 50)

                # Clean markdown fences if present
                cleaned = raw_result
                if "```json" in cleaned:
                    cleaned = cleaned.split("```json")[1].split("```")[0].strip()
                elif "```" in cleaned:
                    cleaned = cleaned.split("```")[1].split("```")[0].strip()

                try:
                    parsed_data = json.loads(cleaned)

                    if current_task == "core_strategy":
                        state["extra_data"]["strategy_plan"]["strategy"] = parsed_data
                    elif current_task == "phases":
                        state["extra_data"]["strategy_plan"]["phases"] = parsed_data
                    elif current_task == "channels":
                        state["extra_data"]["strategy_plan"]["channels"] = parsed_data
                    elif current_task == "execution_plan":
                        state["extra_data"]["strategy_plan"]["execution_plan"] = parsed_data

                    _ok(f"Parsed & merged '{current_task}'", parsed_data)
                    logger.info(f"Successfully parsed JSON for task: {current_task}")

                except Exception as e:
                    _err(f"JSON parse failed for '{current_task}'", e)
                    print(f"  Raw output was:\n{raw_result}")
                    logger.error(f"Failed to parse JSON for {current_task}: {e}")
                    logger.error(f"Raw output: {raw_result}")

            except Exception as e:
                _err("AI Client call failed", e)
                logger.error(f"AI Client execution failed: {e}")

        state["iteration"] = iteration
        state["confidence"] = 0.0
        return state

    # ─────────────────────────────────────────────────────────────
    # STEP 3: REFLECT
    # ─────────────────────────────────────────────────────────────
    async def _reflect(self, state: AgentState) -> AgentState:
        remaining = state.get("plan", [])
        _banner("REFLECT — Evaluating Progress")

        if remaining:
            state["confidence"] = 0.0
            _step(f"Tasks remaining: {len(remaining)}", remaining)
            _step("Decision", "→ Loop back to PLAN (confidence=0.0)")
            logger.info(f"Reflect: {len(remaining)} tasks left, looping back.")
        else:
            state["confidence"] = 0.95
            _ok("All sub-tasks complete! Routing to FINALIZE (confidence=0.95)")
            logger.info("Reflect: All tasks done, routing to finalize.")

        state["needs_human_review"] = False
        return state

    # ─────────────────────────────────────────────────────────────
    # STEP 4: FINALIZE
    # ─────────────────────────────────────────────────────────────
    async def _finalize(self, state: AgentState) -> AgentState:
        strategy = state["extra_data"]["strategy_plan"]

        _banner("FINALIZE — Building Structured JSON Output")

        # ── Structured JSON plan ──────────────────────────────────
        final_plan = {
            "company_name":    strategy["company_name"],
            "industry":        strategy["industry"],
            "target_audience": strategy["target_audience"],
            "generated_at":    strategy["created_at"],
            "strategy": {
                "business_goal":  strategy["strategy"].get("business_goal", ""),
                "marketing_goal": strategy["strategy"].get("marketing_goal", ""),
                "core_strategy":  strategy["strategy"].get("core_strategy", [])
            },
            "campaign_plans": {
                "phases": [
                    {
                        "phase":    phase.get("phase", ""),
                        "duration": phase.get("duration", ""),
                        "tasks":    phase.get("tasks", []) if isinstance(phase.get("tasks"), list)
                                    else [str(phase.get("tasks", ""))]
                    }
                    for phase in strategy.get("phases", [])
                ],
                "channels": [
                    {
                        "channel":   channel.get("channel", ""),
                        "objective": channel.get("objective", ""),
                        "kpi":       channel.get("kpi", "")
                    }
                    for channel in strategy.get("channels", [])
                ],
                "execution_plan": [
                    {
                        "week":     item.get("week", ""),
                        "activity": item.get("activity", ""),
                        "owner":    item.get("owner", ""),
                        "status":   item.get("status", "Pending")
                    }
                    for item in strategy.get("execution_plan", [])
                ]
            }
        }

        state["output"] = final_plan

        print("\n  📋 FINAL STRATEGY PLAN JSON OUTPUT:")
        print("  " + "═" * 58)
        print(json.dumps(final_plan, indent=4, ensure_ascii=False))
        print("  " + "═" * 58)
        _ok(f"Plan generated for: {strategy['company_name']}")
        _step(f"Phases:        {len(final_plan['campaign_plans']['phases'])}")
        _step(f"Channels:      {len(final_plan['campaign_plans']['channels'])}")
        _step(f"Exec Timeline: {len(final_plan['campaign_plans']['execution_plan'])} weeks")

        # ── Persist to database ───────────────────────────────────
        _banner("FINALIZE — Saving Plan to Database")
        _step("tenant_id",    state["tenant_id"])
        _step("company_name", strategy["company_name"])
        _step("user_prompt",  state["extra_data"].get("user_prompt", "")[:80])
        try:
            import traceback
            async with async_session_maker() as session:
                db_plan = StrategyPlan(
                    tenant_id=state["tenant_id"],
                    project_id=state.get("project_id"),
                    product_id=state["extra_data"].get("product_id"),
                    company_name=strategy["company_name"],
                    industry=strategy["industry"],
                    user_prompt=state["extra_data"].get("user_prompt", ""),
                    plan_json=json.dumps(final_plan, ensure_ascii=False)
                )
                session.add(db_plan)
                await session.commit()
                await session.refresh(db_plan)
            _ok(f"Plan saved to DB with ID: {db_plan.id}")
            logger.info(f"Saved StrategyPlan to DB: id={db_plan.id}")
        except Exception as e:
            _err("CRITICAL — Failed to save plan to DB", e)
            print(traceback.format_exc())
            logger.error(f"Failed to save StrategyPlan to DB: {e}\n{traceback.format_exc()}")
        _banner("FINALIZE — Generating Excel Report")
        workbook = Workbook()

        overview_sheet = workbook.active
        overview_sheet.title = "Overview"
        for row in [
            ["Company Name", strategy["company_name"]],
            ["Industry",     strategy["industry"]],
            ["Target Audience", strategy["target_audience"]],
            ["Generated At", strategy["created_at"]],
        ]:
            overview_sheet.append(row)

        strategy_sheet = workbook.create_sheet(title="Strategy")
        strategy_sheet.append(["Category", "Details"])
        strategy_sheet.append(["Business Goal",  strategy["strategy"].get("business_goal", "")])
        strategy_sheet.append(["Marketing Goal", strategy["strategy"].get("marketing_goal", "")])
        for item in strategy["strategy"].get("core_strategy", []):
            strategy_sheet.append(["Core Strategy", item])

        phase_sheet = workbook.create_sheet(title="Execution Phases")
        phase_sheet.append(["Phase", "Duration", "Task"])
        for phase in strategy.get("phases", []):
            tasks = phase.get("tasks", [])
            if not isinstance(tasks, list):
                tasks = [str(tasks)]
            for task in tasks:
                phase_sheet.append([phase.get("phase", ""), phase.get("duration", ""), task])

        channel_sheet = workbook.create_sheet(title="Marketing Channels")
        channel_sheet.append(["Channel", "Objective", "KPI"])
        for channel in strategy.get("channels", []):
            channel_sheet.append([channel.get("channel", ""), channel.get("objective", ""), channel.get("kpi", "")])

        execution_sheet = workbook.create_sheet(title="Execution Timeline")
        execution_sheet.append(["Week", "Activity", "Owner", "Status"])
        for item in strategy.get("execution_plan", []):
            execution_sheet.append([item.get("week", ""), item.get("activity", ""), item.get("owner", ""), item.get("status", "")])

        header_fill = PatternFill(start_color="8C1FF9", end_color="8C1FF9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        state["excel_file"] = excel_buffer.getvalue()

        _ok("Excel report generated successfully")
        _banner("✅ STRATEGY AGENT COMPLETE")
        logger.info(f"Strategy plan finalized for company: {strategy['company_name']}")
        return state